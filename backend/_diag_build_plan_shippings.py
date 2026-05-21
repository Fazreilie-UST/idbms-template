"""
Diagnostic: dump what parse_build_plan_column() actually parses for a
specific config_number, so we can see whether the SUM-formula branch is
firing and whether requestor names resolve to users.

Usage:
    cd backend
    source .venv/bin/activate
    python _diag_build_plan_shippings.py PeP2613.3AD
"""
from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from app.scripts.seed_build_plan import (
    parse_build_plan_column,
    _samples_use_sum_format,
    get_user_by_sample_name,
    clean,
)
from app.db.session import SessionLocal

BUILD_PLAN_DIR = Path("/home/fbinalex/NPI-IDBMS/backend/data/build plan")


def find_column_for_config(df: pd.DataFrame, target: str) -> int | None:
    """Locate the column whose Build Info -> Config Number row matches."""
    target_norm = target.strip().lower()
    for col in range(1, df.shape[1]):
        for _, row in df.iterrows():
            key = clean(row.iloc[0])
            val = clean(row.iloc[col])
            if key and key.lower().startswith("config") and val and val.lower() == target_norm:
                return col
    return None


def main(config_number: str):
    files = sorted(BUILD_PLAN_DIR.glob("*.xlsx"))
    if not files:
        print(f"No xlsx files in {BUILD_PLAN_DIR}")
        return

    for file_path in files:
        print(f"\n=== {file_path.name} ===")
        try:
            xls = pd.ExcelFile(file_path, engine="calamine")
        except Exception as exc:
            print(f"  could not open: {exc}")
            continue

        wb = load_workbook(file_path, data_only=True, read_only=False)
        wb_formula = load_workbook(file_path, data_only=False, read_only=False)

        for sheet_name in xls.sheet_names:
            if sheet_name == "Shipping Info":
                continue
            df = pd.read_excel(
                file_path, engine="calamine", sheet_name=sheet_name, header=None
            ).dropna(how="all").dropna(axis=1, how="all")

            col = find_column_for_config(df, config_number)
            if col is None:
                continue

            print(f"  FOUND in sheet={sheet_name!r}, column={col}")
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else None
            ws_formula = wb_formula[sheet_name] if sheet_name in wb_formula.sheetnames else None

            parsed = parse_build_plan_column(df, col, ws=ws, ws_formula=ws_formula)
            samples = parsed["samples"]

            print(f"  samples count: {len(samples)}")
            print(f"  _samples_use_sum_format -> {_samples_use_sum_format(samples)}")
            print(f"\n  --- sample rows ---")
            for s in samples:
                print(
                    f"    row_idx={s['row_idx']:>3}  "
                    f"bold={int(s['is_bold'])}  "
                    f"formula_refs={s['formula_refs']}  "
                    f"name={s['name']!r}  value={s['value']!r}"
                )

            # Check which names resolve to users
            print(f"\n  --- user resolution ---")
            with SessionLocal() as session:
                for s in samples:
                    name = clean(s["name"])
                    if not name:
                        continue
                    user = get_user_by_sample_name(session, name)
                    flag = "OK" if user else "MISS"
                    uid = user.id if user else "-"
                    print(f"    [{flag}] {name!r:50s} -> user_id={uid}")
            return

    print(f"\nConfig number {config_number!r} not found in any sheet/file.")


if __name__ == "__main__":
    cfg = sys.argv[1] if len(sys.argv) > 1 else "PeP2613.3AD"
    main(cfg)
