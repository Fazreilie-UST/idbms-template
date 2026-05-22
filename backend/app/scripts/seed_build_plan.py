# app/scripts/seed_build_plan.py

import logging
import math
from difflib import SequenceMatcher
from pathlib import Path
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from sqlalchemy.orm import Session

from app.db.session import SessionLocal
from app.models.auth.user import User

from app.models.build.family import Family
from app.models.build.form_factor import FormFactor
from app.models.build.family_form_factor import FamilyFormFactor

from app.models.build.build_plan import (
    BuildPlan,
    BuildPlanBuildDesc,
    BuildPlanBuildNote,
    BuildNote,
    SupportActivity,
    SupportActivityBuildNote,
    BuildPlanStatus,
)

from app.models.build.build_plan_component import BuildPlanComponent
from app.models.build.component import Component, ComponentSlot
from app.models.build.supplier import Supplier
from app.models.build.component_supplier import ComponentSupplier
from app.models.build.component_supplier_family import ComponentSupplierFamily
from app.models.build.attribute_definition import AttributeDefinition
from app.models.build.silicon_stepping import (
    SiliconStepping,
    BuildPlanSiliconStepping,
    BASE_SILICON_STEPPINGS,
)
from app.models.build.component_attribute_value import ComponentAttributeValue

from app.models.build.warehouse import Warehouse, QuantityStoredInWarehouse
from app.models.build.build_plan_build_request import BuildPlanBuildRequest
from app.models.build.test import Test, TestDetail
from app.models.build.build_plan_test import BuildPlanTest
from app.models.build.config_number import ConfigNumber
from app.models.build.build_plan_access import BuildPlanAccess, AccessTypeEnum
from app.models.build.build_plan_shipping import BuildPlanShipping

from app.models.order.build_request import BuildRequest, BuildRequestStatus

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Family name -> code mapping
# ---------------------------------------------------------------------------
# Spreadsheets sometimes contain the long-form family name (e.g.
# "Whale Peak") in the Family row. We canonicalise these to a short code
# (e.g. "WhP") so downstream lookups, filters and dashboards work against a
# stable identifier. Add new mappings here as new families appear.
#
# Keys are matched case-insensitively after collapsing whitespace. The
# trailing "2" suffix that appears on workbook family labels (e.g.
# ``"Whale Peak 2"`` or family code ``"WhP2"``) is intentionally collapsed
# to the base family (``"WhP"`` / ``"Whale Peak"``) at parse time so that
# both editions of a family share a single row in the database.
FAMILY_NAME_TO_CODE: dict[str, tuple[str, str]] = {
    # parsed-name (lower-case)  -> (canonical_code, canonical_name)
    "whale peak":      ("WhP",  "Whale Peak"),
    "lizard peak":     ("LzP",  "Lizard Peak"),
    "pelican peak":    ("PeP",  "Pelican Peak"),
    "spider peak":     ("SpP",  "Spider Peak"),
}

# Regex used to strip a trailing ``" 2"`` / ``"2"`` suffix from family
# labels (both long-form names and short codes). Examples:
#   ``"Whale Peak 2"`` -> ``"Whale Peak"``
#   ``"WhP2"``         -> ``"WhP"``
#   ``"LzP2"``         -> ``"LzP"``
_FAMILY_TRAILING_TWO = re.compile(r"\s*2\s*$")


def resolve_family_identity(raw_family: str) -> tuple[str, str]:
    """Return ``(canonical_code, canonical_name)`` for a parsed family string.

    Behaviour
    ---------
    * A trailing ``2`` suffix on the parsed label (``"Whale Peak 2"`` or
      a bare code ``"WhP2"``) is stripped so the two editions of a family
      share one row.
    * Lookup is case-insensitive with collapsed whitespace; any value not
      in :data:`FAMILY_NAME_TO_CODE` is returned verbatim (used as both
      code and name) so newly-seen families still produce a workable row.
    """
    cleaned = (raw_family or "").strip()
    if not cleaned:
        return "", ""

    # Strip trailing " 2" / "2" suffix.
    cleaned = _FAMILY_TRAILING_TWO.sub("", cleaned).strip() or cleaned

    key = " ".join(cleaned.split()).lower()
    mapping = FAMILY_NAME_TO_CODE.get(key)
    if mapping is not None:
        return mapping
    return cleaned, cleaned


# ---------------------------------------------------------------------------
# Supplier name normalisation
# ---------------------------------------------------------------------------
# Spreadsheets often carry the same vendor under a handful of inconsistent
# spellings (case differences, typos, trailing whitespace). We collapse
# those to a single canonical row in ``suppliers`` so downstream filters,
# dashboards and the admin Components/Suppliers tree stay clean.
#
# Keys are matched case-insensitively after collapsing whitespace. Values
# are the canonical display name. Extend this dict whenever a new
# duplicate is spotted in the wild.
SUPPLIER_ALIASES: dict[str, str] = {
    # speedtech family
    "speedtech":  "SpeedTech",
    "speed tech": "SpeedTech",
    "sppedtech":  "SpeedTech",
    "speedteck":  "SpeedTech",
}

# Minimum SequenceMatcher ratio for two supplier names to be treated as
# duplicates of one another when no explicit alias exists. 0.85 picks up
# single-character typos in ~6-12 char vendor names without merging
# legitimately distinct vendors.
SUPPLIER_SIMILARITY_THRESHOLD = 0.85

_SUPPLIER_WHITESPACE = re.compile(r"\s+")


def _supplier_key(name: str) -> str:
    """Lower-case, whitespace-collapsed lookup key for supplier names."""
    return _SUPPLIER_WHITESPACE.sub(" ", (name or "").strip()).lower()


def canonicalize_supplier_name(raw: str) -> str:
    """Return the canonical display name for ``raw`` using the alias map.

    Unknown names are returned with surrounding/duplicate whitespace
    stripped; the alias map is consulted case-insensitively.
    """
    cleaned = _SUPPLIER_WHITESPACE.sub(" ", (raw or "").strip())
    if not cleaned:
        return ""
    return SUPPLIER_ALIASES.get(cleaned.lower(), cleaned)


def resolve_supplier(session: Session, raw_name: str) -> Supplier | None:
    """Return a :class:`Supplier` row for ``raw_name``, merging near-duplicates.

    Resolution order:

    1. Apply :data:`SUPPLIER_ALIASES` to canonicalise the input.
    2. Exact (case-insensitive) match against existing supplier rows.
    3. Fuzzy match (``SequenceMatcher`` ratio ≥
       :data:`SUPPLIER_SIMILARITY_THRESHOLD`) against existing rows; a
       warning is logged so operators can decide whether to add the
       observed spelling to :data:`SUPPLIER_ALIASES`.
    4. Insert a new row using the canonical name.
    """
    from sqlalchemy import func as _sa_func  # local to avoid top-level cost

    cleaned = (raw_name or "").strip()
    if not cleaned:
        return None

    canonical = canonicalize_supplier_name(cleaned)
    key = canonical.lower()

    # Stage 1+2: exact match on the canonical form.
    row = (
        session.query(Supplier)
        .filter(_sa_func.lower(Supplier.name) == key)
        .first()
    )
    if row:
        return row

    # Stage 3: fuzzy match against everything currently in the DB.
    best: Supplier | None = None
    best_ratio = 0.0
    for cand in session.query(Supplier).all():
        ratio = SequenceMatcher(None, key, (cand.name or "").lower()).ratio()
        if ratio > best_ratio:
            best_ratio = ratio
            best = cand
    if best is not None and best_ratio >= SUPPLIER_SIMILARITY_THRESHOLD:
        logger.warning(
            "Supplier name %r matched existing %r (ratio %.2f); reusing. "
            "Consider adding %r -> %r to SUPPLIER_ALIASES.",
            cleaned, best.name, best_ratio, cleaned.lower(), best.name,
        )
        return best

    # Stage 4: brand-new vendor.
    row = Supplier(name=canonical)
    session.add(row)
    session.flush()
    return row


def merge_duplicate_suppliers(
    session: Session,
    aliases: dict[str, str] | None = None,
) -> dict[str, int | dict[str, int]]:
    """Repoint every alias supplier in ``aliases`` to its canonical row and
    delete the duplicates.

    The function moves three relationships to the canonical supplier:

      * ``build_plan_components.supplier_id``
      * ``component_suppliers``  (idempotent: drops rows that would
        violate the composite PK)
      * ``component_supplier_families`` (same)

    Returns a stats dict suitable for logging::

        {
          "merged_pairs": int,
          "deleted_suppliers": int,
          "repointed": {
            "build_plan_components": int,
            "component_suppliers": int,
            "component_supplier_families": int,
          },
        }

    Safe to re-run; if the canonical supplier doesn't exist yet it will be
    created (so the function also doubles as "ensure the SpeedTech-style
    canonical rows are in the DB").
    """
    alias_map = aliases or SUPPLIER_ALIASES
    stats: dict = {
        "merged_pairs": 0,
        "deleted_suppliers": 0,
        "repointed": {
            "build_plan_components": 0,
            "component_suppliers": 0,
            "component_supplier_families": 0,
        },
    }

    # Group alias rows by canonical name so multiple typos collapse into
    # the same target in one pass.
    by_canonical: dict[str, list[str]] = {}
    for raw_alias, canonical in alias_map.items():
        by_canonical.setdefault(canonical, []).append(raw_alias)

    for canonical_name, alias_keys in by_canonical.items():
        canonical = (
            session.query(Supplier)
            .filter(Supplier.name == canonical_name)
            .first()
        )
        if canonical is None:
            canonical = Supplier(name=canonical_name)
            session.add(canonical)
            session.flush()

        # Find every supplier row whose name matches any of the aliases
        # (case-insensitive) AND isn't the canonical itself.
        alias_lower = [a.lower() for a in alias_keys]
        dup_rows = (
            session.query(Supplier)
            .filter(
                Supplier.id != canonical.id,
                Supplier.name.ilike("%"),  # placeholder; refined below
            )
            .all()
        )
        # Cheap Python-side filter (handles whitespace + arbitrary case).
        dup_rows = [
            r for r in dup_rows if _supplier_key(r.name) in alias_lower
        ]

        for dup in dup_rows:
            # -- 1. build_plan_components.supplier_id ---------------------
            from app.models.build.build_plan_component import BuildPlanComponent as _BPC
            n = (
                session.query(_BPC)
                .filter(_BPC.supplier_id == dup.id)
                .update({"supplier_id": canonical.id}, synchronize_session=False)
            )
            stats["repointed"]["build_plan_components"] += int(n or 0)

            # -- 2. component_supplier_families ----------------------------
            # Move all CSF rows; drop those that would clash with an
            # existing (component, canonical, family) row.
            csf_rows = (
                session.query(ComponentSupplierFamily)
                .filter(ComponentSupplierFamily.supplier_id == dup.id)
                .all()
            )
            for csf in csf_rows:
                clash = (
                    session.query(ComponentSupplierFamily)
                    .filter(
                        ComponentSupplierFamily.component_id == csf.component_id,
                        ComponentSupplierFamily.supplier_id == canonical.id,
                        ComponentSupplierFamily.family_id == csf.family_id,
                    )
                    .first()
                )
                if clash:
                    session.delete(csf)
                else:
                    csf.supplier_id = canonical.id
                    stats["repointed"]["component_supplier_families"] += 1
            session.flush()

            # -- 3. component_suppliers -----------------------------------
            cs_rows = (
                session.query(ComponentSupplier)
                .filter(ComponentSupplier.supplier_id == dup.id)
                .all()
            )
            for cs in cs_rows:
                clash = (
                    session.query(ComponentSupplier)
                    .filter(
                        ComponentSupplier.component_id == cs.component_id,
                        ComponentSupplier.supplier_id == canonical.id,
                    )
                    .first()
                )
                if clash:
                    session.delete(cs)
                else:
                    cs.supplier_id = canonical.id
                    stats["repointed"]["component_suppliers"] += 1
            session.flush()

            session.delete(dup)
            stats["merged_pairs"] += 1
            stats["deleted_suppliers"] += 1

    session.flush()
    return stats


# ---------------------------------------------------------------------------
# Year derivation from config number
# ---------------------------------------------------------------------------
# Config numbers follow the shape "<FamilyCode><YY><WW>" (e.g. "WhP2618"
# = year 2026, work-week 18). The family code is variable length and may
# include digits, so we anchor on the first 4-digit run that immediately
# follows the alphabetic family-code prefix. Anything after (e.g. ``.1AD``,
# ``.1 rQS``) is ignored.
_CONFIG_YEAR_WW_PATTERN = re.compile(r"[A-Za-z](\d{2})(\d{2})")


def derive_year_from_config(config_number: str | None) -> int | None:
    """Return the 4-digit calendar year encoded in a config number, or
    ``None`` when the format doesn't match. ``"WhP2618"`` -> ``2026``.
    """
    value = clean(config_number) if config_number is not None else None
    if not value:
        return None
    match = _CONFIG_YEAR_WW_PATTERN.search(value)
    if not match:
        return None
    return 2000 + int(match.group(1))


def derive_work_week_from_config(config_number: str | None) -> int | None:
    """Return the ISO work-week (1-53) encoded in a config number, or
    ``None`` when the format doesn't match. ``"WhP2618"`` -> ``18``.
    """
    value = clean(config_number) if config_number is not None else None
    if not value:
        return None
    match = _CONFIG_YEAR_WW_PATTERN.search(value)
    if not match:
        return None
    ww = int(match.group(2))
    if ww < 1 or ww > 53:
        return None
    return ww


SECTION_NAMES = {
    "Build Info",
    "Key Components",
    "Test Section",
    "Quantities",
    "Samples",
}

IGNORE_ROWS = {"total"}

# Free-form labels seen in the Excel sheets in the warehouse / quantity-storage
# rows. Each label is parsed as "<warehouse code><suffix>", where the suffix
# is optional and may describe a sub-bucket ("KEEP FOR TEST", "TO KEEP FOR
# TEST - Solder Down", ...). The importer collapses all variants to the
# canonical warehouse name on the right.
WAREHOUSE_CODES = ("CNB5", "AZW", "ODM")

# Regex: warehouse-code prefix followed by an optional space/punctuation
# delimited descriptor. The descriptor is captured but ignored — any text
# after the code is treated as a sub-bucket of that warehouse.
_WAREHOUSE_ALIAS_PATTERN = re.compile(
    r"^\s*(?P<code>" + "|".join(WAREHOUSE_CODES) + r")\b(?P<rest>.*)$",
    re.IGNORECASE,
)


def resolve_warehouse_alias(value) -> str | None:
    """Return the canonical warehouse name (``AZW`` / ``CNB5`` / ``ODM``)
    for a free-form Excel label, or ``None`` if the label is not a
    warehouse-stash row.

    Examples (all → ``"AZW"``)::

        "AZW"
        "AZW KEEP FOR TEST"
        "AZW TO KEEP FOR TEST"
        "AZW TO KEEP FOR TEST - Solder Down"
    """
    if value is None:
        return None
    text_value = str(value).strip()
    if not text_value:
        return None
    match = _WAREHOUSE_ALIAS_PATTERN.match(text_value)
    if not match:
        return None
    return match.group("code").upper()

VALID_COMPONENTS = {
    "PCB",
    "ADP PCB",
    "Diplexer",
    "Crystal",
    "Inductor",
    "RF_Connector",
    "Cap_3T",
    "Ferrite_Bead",
    "Shield",
    "Silicon",
    "HW",
}

VALID_SLOTS = {"Ch_A", "Ch_B", "1", "2"}

SPACE_FIELD_MAP = {
    "PCB Supplier": ("PCB", None, "Supplier"),
    "PCB Revision": ("PCB", None, "Revision"),
    "HW Revision": ("HW", None, "Revision"),
    "ADP PCB Supplier": ("ADP PCB", None, "Supplier"),
    "ADP PCB Version": ("ADP PCB", None, "Version"),
    "Shield Supplier": ("Shield", None, "Supplier"),
    "Shield Version": ("Shield", None, "Version"),
    # "Silicon Stepping" is intentionally NOT routed through the component
    # attribute path; it is split on whitespace and stored in the
    # ``silicon_steppings`` many-to-many table (see
    # :func:`import_silicon_steppings`).
    "Silicon Package Version": ("Silicon", None, "Package Version"),
    "Silicon Test Program": ("Silicon", None, "Test Program"),
}

# Field name (in Key Components) that carries the build plan's silicon
# stepping values. Stored separately from component attributes.
SILICON_STEPPING_FIELD = "Silicon Stepping"

SAMPLE_GROUP_NAMES = {"IDC", "PTK"}

OWNER_EMPLOYEE_ID = "10692120"  # Choi, Wai Mee
EDITOR_EMPLOYEE_IDS = {"12282943"}  # Bin Alep. Fazreilie

ACCESS_LEVEL_RANK = {
    AccessTypeEnum.editor: 1,
    AccessTypeEnum.owner: 2,
}


def clean(value):
    if pd.isna(value):
        return None
    value = str(value).strip()
    return value or None


def normalize_text(value: str):
    value = clean(value)
    if not value:
        return None
    value = re.sub(r"\s+", " ", value)
    return value.strip().lower()


def safe_int(value, default=None):
    value = clean(value)
    if value is None:
        return default

    try:
        return int(float(value))
    except ValueError:
        return default


def safe_int_ceil(value, default=None):
    """Like :func:`safe_int` but rounds up. Used for quantities derived from
    Excel formulas (e.g. ``Build Start Quantity = required / yield``) so
    fractional results cover all required units (33 / 0.90 -> 37)."""
    value = clean(value)
    if value is None:
        return default

    try:
        return math.ceil(float(value))
    except ValueError:
        return default
    

def safe_percent_int(value, default=None):
    value = clean(value)
    if value is None:
        return default

    value = value.replace("%", "").strip()

    try:
        number = float(value)
    except ValueError:
        return default

    # Excel percentage cell: 90% becomes 0.9
    if 0 < number <= 1:
        number = number * 100

    return int(round(number))


def get_or_create(session: Session, model, defaults=None, **kwargs):
    obj = session.query(model).filter_by(**kwargs).first()
    if obj:
        return obj

    params = dict(kwargs)
    if defaults:
        params.update(defaults)

    obj = model(**params)
    session.add(obj)
    session.flush()
    return obj


def is_others_row(value: str):
    value = normalize_text(value)
    if not value:
        return False

    return (
        value.startswith("others")
        and "email" in value
        and "instruction" in value
    )


def is_sample_group_header(value: str):
    value = clean(value)
    if not value:
        return False

    return value.upper() in SAMPLE_GROUP_NAMES


def is_non_user_sample_row(value: str):
    """Return True if the cell is a warehouse-stash row (any AZW/ODM/CNB5
    variant, including suffixes like ``"AZW TO KEEP FOR TEST - Solder Down"``)
    or otherwise not a user name.
    """
    cleaned = clean(value)
    if not cleaned:
        return True
    return resolve_warehouse_alias(cleaned) is not None


def map_build_status(value):
    value = clean(value)
    if not value:
        return BuildPlanStatus.new

    # Spreadsheet aliases that mean "no real status yet" -> treat as `new`,
    # matching the behaviour of an empty cell.
    EMPTY_ALIASES = {"placeholder", "tbd", "tba", "n/a", "na", "none", "-"}
    if value.lower() in EMPTY_ALIASES:
        return BuildPlanStatus.new

    for enum_value in BuildPlanStatus:
        if enum_value.value.lower() == value.lower():
            return enum_value

    # Other common phrasings that don't map 1:1 to the enum.
    aliases = {
        "pending": BuildPlanStatus.hold,
        "on hold": BuildPlanStatus.hold,
        "in progress": BuildPlanStatus.plan,
        "completed": BuildPlanStatus.done,
        "complete": BuildPlanStatus.done,
        "finished": BuildPlanStatus.done,
        "cancel": BuildPlanStatus.cancelled,
        "canceled": BuildPlanStatus.cancelled,
        "dropped": BuildPlanStatus.cancelled,
    }
    mapped = aliases.get(value.lower())
    if mapped is not None:
        return mapped

    # Truly unknown values shouldn't blow up the whole import; default to
    # `new` so a human can review the row later.
    return BuildPlanStatus.new


def _is_cell_bold(ws, row_1based: int, col_1based: int) -> bool:
    """Read the bold flag for a worksheet cell. Returns False on any
    error or missing font info so that classification falls back to the
    "non-bold" (member) path instead of the "handler" path.
    """
    if ws is None:
        return False
    try:
        cell = ws.cell(row=row_1based, column=col_1based)
        font = cell.font
        return bool(font and font.bold)
    except Exception:
        return False


# Matches a SUM-only formula (case-insensitive). We deliberately do NOT
# support arithmetic chains like ``=B12+B13`` so the SUM-format detection
# stays unambiguous. Build-plan sheets that use the "handler holds the
# total" pattern always use SUM(...) according to the PMs maintaining
# them.
_SUM_FORMULA_PATTERN = re.compile(r"^\s*=\s*SUM\s*\((.*)\)\s*$", re.IGNORECASE | re.DOTALL)
_CELL_REF_PATTERN = re.compile(r"^\$?([A-Z]+)\$?(\d+)$", re.IGNORECASE)


def _parse_sum_formula_rows(formula) -> set[int] | None:
    """Return the set of *0-based dataframe row indices* referenced by a
    ``=SUM(...)`` formula, or ``None`` if ``formula`` is not a SUM-only
    formula.

    Handles ranges (``SUM(B12:B15)``) and explicit comma-separated refs
    (``SUM(B12, B14, B15)``). Mixed forms (``SUM(B12:B14, B16)``) are
    supported too. Cell-column letters are ignored \u2014 only the row part
    matters because the importer scans one column at a time and any
    formula it sees should reference rows in that same column.
    """
    if formula is None:
        return None
    if not isinstance(formula, str):
        return None
    match = _SUM_FORMULA_PATTERN.match(formula)
    if not match:
        return None

    body = match.group(1)
    rows: set[int] = set()

    for piece in body.split(","):
        piece = piece.strip()
        if not piece:
            continue

        if ":" in piece:
            try:
                _min_col, min_row, _max_col, max_row = range_boundaries(piece)
            except (ValueError, TypeError):
                continue
            if min_row is None or max_row is None:
                continue
            for r in range(min_row, max_row + 1):
                rows.add(r - 1)
            continue

        cell_match = _CELL_REF_PATTERN.match(piece)
        if cell_match:
            rows.add(int(cell_match.group(2)) - 1)

    return rows or None


def parse_build_plan_column(df: pd.DataFrame, col: int, ws=None, ws_formula=None):
    parsed = {
        "build_info": {},
        "key_components": {},
        "test_section": {},
        "quantities": {},
        # Samples are stored as dicts so we can carry per-row metadata
        # (bold flag, formula references, source row index) alongside
        # the basic ``(name, value)`` pair. Older callers still expect
        # tuple-like access \u2014 see :func:`_iter_sample_rows`.
        "samples": [],
        "warehouse_quantities": {},
    }

    current_section = None

    # Column index in the original (pre-dropna) sheet so we can look up
    # formatting via openpyxl. ``df.columns`` keeps the original integer
    # labels even after ``dropna(axis=1)``.
    try:
        original_col_index = int(df.columns[col])
    except Exception:
        original_col_index = col
    openpyxl_col = original_col_index + 1

    for orig_idx, row in df.iterrows():
        key = clean(row.iloc[0])
        value = clean(row.iloc[col])

        if not key:
            continue

        if key in SECTION_NAMES:
            current_section = key
            continue

        if key.strip().lower() in IGNORE_ROWS:
            continue

        canonical_warehouse = resolve_warehouse_alias(key)
        if canonical_warehouse:
            parsed["warehouse_quantities"][canonical_warehouse] = value
            continue

        if current_section == "Build Info":
            parsed["build_info"][key] = value

        elif current_section == "Key Components":
            parsed["key_components"][key] = value

        elif current_section == "Test Section":
            parsed["test_section"][key] = value

        elif current_section == "Quantities":
            parsed["quantities"][key] = value

        elif current_section == "Samples":
            try:
                openpyxl_row = int(orig_idx) + 1
            except Exception:
                openpyxl_row = None

            # The requester name lives in column 0 of the same row; use
            # that cell's bold flag (not the value cell's) as the group
            # signal, since the spec is about the requester text style.
            is_bold = (
                _is_cell_bold(ws, openpyxl_row, 1)
                if openpyxl_row is not None
                else False
            )

            # Read the raw formula from the value cell (if any). Requires
            # the workbook to be opened with ``data_only=False``; otherwise
            # ``cell.value`` returns the cached numeric result and the
            # SUM detection silently falls back to the "Others/bold"
            # format. ``ws_formula`` is passed in by the caller.
            formula_refs = None
            if ws_formula is not None and openpyxl_row is not None:
                try:
                    raw = ws_formula.cell(row=openpyxl_row, column=openpyxl_col).value
                except Exception:
                    raw = None
                formula_refs = _parse_sum_formula_rows(raw)

            try:
                df_row_idx = int(orig_idx)
            except Exception:
                df_row_idx = -1

            parsed["samples"].append({
                "name": key,
                "value": value,
                "is_bold": is_bold,
                "formula_refs": formula_refs,
                "row_idx": df_row_idx,
            })

    return parsed


def parse_component_field(field_name: str):
    name = field_name.strip()

    if name in SPACE_FIELD_MAP:
        return SPACE_FIELD_MAP[name]

    parts = name.split("_")

    if len(parts) >= 4:
        component_name = parts[0]
        slot_code = f"{parts[1]}_{parts[2]}"
        attr = "_".join(parts[3:])
        return component_name, slot_code, attr

    if name.startswith("RF_Connector_"):
        return "RF_Connector", parts[-1], "Supplier"

    if name.startswith("Ferrite_Bead_"):
        return "Ferrite_Bead", parts[-1], "Supplier"

    if name.startswith("Cap_3T_"):
        return "Cap_3T", parts[-1], "Supplier"

    if len(parts) == 2:
        return parts[0], parts[1], "Supplier"

    raise ValueError(f"Unmapped component field: {field_name}")


def get_family_form_factor(
    session: Session,
    family_code: str,
    form_factor_name: str,
    family_name: str | None = None,
):
    """Upsert a (Family, FormFactor, FamilyFormFactor) triple and return the
    join-row. The FormFactor table only stores ``name`` (the canonical label
    such as ``"1216 Module"`` or ``"1216 Adaptor Module"``).
    """
    # ``family_name`` defaults to ``family_code`` for callers that haven't
    # been updated yet (keeps backward compatibility with older invocations).
    canonical_name = family_name or family_code

    family = get_or_create(
        session,
        Family,
        code=family_code,
        defaults={"name": canonical_name},
    )

    # Backfill the human-readable name if the row was created earlier with
    # only the code (e.g. before the FAMILY_NAME_TO_CODE mapping existed).
    if canonical_name and family.name != canonical_name:
        # Only overwrite a placeholder where name == code; never clobber a
        # user-edited display name.
        if family.name == family.code:
            family.name = canonical_name
            session.flush()

    # FormFactors are unique on ``name`` (the full canonical label, e.g.
    # ``"1216 Adaptor Module"``).
    form_factor = get_or_create(
        session,
        FormFactor,
        name=form_factor_name,
    )
    session.flush()

    family_form_factor = (
        session.query(FamilyFormFactor)
        .filter_by(family_id=family.id, form_factor_id=form_factor.id)
        .first()
    )

    if not family_form_factor:
        family_form_factor = FamilyFormFactor(
            family_id=family.id, form_factor_id=form_factor.id
        )
        session.add(family_form_factor)
        session.flush()

    return family_form_factor


# Recognised trailing variant tokens (case-insensitive). ``Adapter`` is a
# synonym of ``Adaptor``; ``Module`` is the default packaging.
_VARIANT_SUFFIX_TOKENS: set[str] = {"module", "adaptor", "adapter"}

# Noise tokens that should be silently dropped during FormFactor normalisation.
# (e.g. "1216 on adaptor" -> "1216 adaptor")
_SKU_NOISE_TOKENS: set[str] = {"on"}

# Recognised qualifier tokens that may sit between the bare code and the
# variant suffix. Order in the canonical name follows the source-of-truth
# casing in this map's values. ``MS`` is intentionally dropped from the
# canonical name (``"2230 MS Module"`` collapses to ``"2230 Module"``) per
# the simplified Form Factor scheme.
_SKU_QUALIFIER_TOKENS: dict[str, str] = {
    "ms": "MS",
    "lte": "LTE",
}

# Qualifiers that are accepted in the parsed source but stripped from the
# canonical Form Factor name.
_SKU_DROP_QUALIFIERS: set[str] = {"MS"}


def _digits_only(token: str) -> str:
    """Strip everything except digits from ``token``."""
    return re.sub(r"\D", "", token)


def _normalise_form_factor_name(raw: str) -> str:
    """Normalise a raw Form Factor label into its canonical name.

    Rules
    -----
    * Standard variants ``<code>`` / ``<code> Module`` / ``<code> MS Module``
      all collapse to ``"<code> Module"``.
    * Adaptor variants ``<code> adaptor`` / ``<code> on adaptor`` /
      ``<code> Adaptor Module`` collapse to ``"<code> Adaptor Module"``.
    * ``<code> LTE Adaptor`` -> ``"<code> LTE Adaptor Module"``.
    * Noise tokens (``"on"``) and the qualifier ``MS`` are dropped from
      the canonical name.
    * The canonical name always ends with ``Module``.

    Examples
    --------
    ``"2230"``                  -> ``"2230 Module"``
    ``"2230 Module"``           -> ``"2230 Module"``
    ``"2230 MS Module"``        -> ``"2230 Module"``
    ``"1216 Adaptor"``          -> ``"1216 Adaptor Module"``
    ``"1216 on Adaptor"``       -> ``"1216 Adaptor Module"``
    ``"1216 Adaptor Module"``   -> ``"1216 Adaptor Module"``
    ``"1216 LTE Adaptor"``      -> ``"1216 LTE Adaptor Module"``
    """
    cleaned = (raw or "").strip()
    if not cleaned:
        return ""

    # Tokenize and drop noise tokens.
    tokens = [
        t for t in cleaned.split()
        if t.lower() not in _SKU_NOISE_TOKENS
    ]
    if not tokens:
        return cleaned

    bare_code = _digits_only(tokens[0]) or tokens[0]
    rest = list(tokens[1:])

    # Strip trailing variant suffix tokens (Module / Adaptor / Adapter).
    suffix_lower: list[str] = []
    while rest and rest[-1].lower() in _VARIANT_SUFFIX_TOKENS:
        suffix_lower.insert(0, rest.pop().lower())

    # Validate that everything left between code and suffix is a known
    # qualifier; otherwise fall back to the raw label.
    qualifier_canonical: list[str] = []
    for token in rest:
        canonical = _SKU_QUALIFIER_TOKENS.get(token.lower())
        if canonical is None:
            return cleaned
        qualifier_canonical.append(canonical)

    # Drop "MS" qualifier from canonical name.
    qualifier_canonical = [q for q in qualifier_canonical if q not in _SKU_DROP_QUALIFIERS]

    has_adaptor = "adaptor" in suffix_lower or "adapter" in suffix_lower

    name_parts: list[str] = [bare_code]
    if qualifier_canonical:
        name_parts.append(" ".join(qualifier_canonical))
    if has_adaptor:
        name_parts.append("Adaptor")
    name_parts.append("Module")
    return " ".join(name_parts)


def detect_family_and_form_factor(df: pd.DataFrame):
    """Parse the Family + Form Factor rows out of a build-plan worksheet.

    Returns ``(family_code, family_name, form_factor_name)``.
    """
    first_col = df.iloc[:, 0].astype(str)

    family_rows = first_col[first_col.str.contains("Family", case=False, na=False)]
    # Accept either the legacy "FormFactor:" label or the newer "SKU:" label
    # (newer build-plan templates use "SKU: <name>" in place of
    # "FormFactor: <name>"). Match as a label prefix to avoid false hits on
    # data rows that happen to mention "sku" somewhere in their text.
    form_factor_rows = first_col[
        first_col.str.contains(r"^\s*(?:FormFactor|SKU)\s*:", case=False, na=False, regex=True)
    ]

    if family_rows.empty:
        raise ValueError("Family row not found")

    if form_factor_rows.empty:
        raise ValueError("FormFactor row not found")

    raw_family = family_rows.iloc[0].split(":", 1)[1].strip()
    raw_form_factor = form_factor_rows.iloc[0].split(":", 1)[1].strip()

    family_code, family_name = resolve_family_identity(raw_family)
    form_factor_name = _normalise_form_factor_name(raw_form_factor)
    return family_code, family_name, form_factor_name


def seed_default_warehouses(session: Session):
    """Ensure the canonical warehouse rows exist.

    The Excel build-plan sheets reference several free-form variants
    (``"AZW KEEP FOR TEST"``, ``"ODM TO KEEP FOR TEST"``, ...) which the
    importer collapses to these canonical names via :func:`resolve_warehouse_alias`.
    Idempotent — safe to call from both the build-plan importer and the
    initial-DB seeding script.
    """
    default_warehouses = [
        {"name": "AZW", "notes": "AZW warehouse (per-build-plan test holding)"},
        {"name": "CNB5", "notes": "CNB5 warehouse"},
        {"name": "ODM", "notes": "ODM warehouse (per-build-plan test holding)"},
    ]

    for item in default_warehouses:
        warehouse = session.query(Warehouse).filter_by(name=item["name"]).first()
        if warehouse:
            if not warehouse.notes:
                warehouse.notes = item["notes"]
        else:
            session.add(Warehouse(name=item["name"], notes=item["notes"]))

    session.flush()


def get_warehouse_map(session: Session):
    return {warehouse.name: warehouse for warehouse in session.query(Warehouse).all()}


def import_warehouse_quantities(
    session: Session,
    build_plan: BuildPlan,
    warehouse_quantities: dict,
):
    warehouse_map = get_warehouse_map(session)

    for source_key, raw_quantity in warehouse_quantities.items():
        # Some build-plan files leave the warehouse-name cell blank; the
        # parser then yields an empty/None key. Skip those rows silently
        # rather than failing the whole import.
        cleaned_key = clean(source_key) if source_key is not None else ""
        if not cleaned_key:
            continue

        warehouse_name = resolve_warehouse_alias(cleaned_key)
        if not warehouse_name:
            continue

        warehouse = warehouse_map.get(warehouse_name)
        if not warehouse:
            raise ValueError(f"Warehouse not found: {warehouse_name}")

        quantity_value = safe_int(raw_quantity, default=0)

        existing = (
            session.query(QuantityStoredInWarehouse)
            .filter_by(buildplan_id=build_plan.id, warehouse_id=warehouse.id)
            .first()
        )

        if existing:
            existing.quantity_stored = quantity_value
        else:
            session.add(
                QuantityStoredInWarehouse(
                    buildplan_id=build_plan.id,
                    warehouse_id=warehouse.id,
                    quantity_stored=quantity_value,
                )
            )


# Threshold above which a sample-row name is considered a match for an
# existing user (via difflib SequenceMatcher.ratio()). Empirically tuned
# so common spreadsheet typos / re-orderings (e.g. "Choi Wai Mee" vs
# "Wai Mee, Choi") collapse to the right user while completely unrelated
# names stay separate.
_USER_FUZZY_MATCH_THRESHOLD = 0.85

# Strip anything inside parentheses (and the parens themselves). PMs often
# annotate sample names like ``"Ali (PTK lead)"`` or ``"Abu (on leave)"``;
# we want to match against the bare user name.
_PARENS_PATTERN = re.compile(r"\([^)]*\)")


def _strip_parens(value: str) -> str:
    if not value:
        return value
    return _PARENS_PATTERN.sub("", value).strip()


def _normalize_for_match(value: str) -> str:
    """Lower-case + collapse whitespace + strip parenthesised content. Used
    only for fuzzy comparison; the user's stored ``full_name`` is never
    mutated."""
    if not value:
        return ""
    cleaned = _strip_parens(str(value)).lower()
    return re.sub(r"\s+", " ", cleaned).strip()


def get_user_by_sample_name(session: Session, requester_name: str):
    """Resolve a sample-row name to a :class:`User` using
    parenthesis-stripping + fuzzy matching.

    Strategy:
      1. Strip parenthesised annotations and try exact (case-insensitive)
         match on ``full_name``.
      2. Fall back to ``difflib.SequenceMatcher`` against every user's
         normalised ``full_name``; return the best match whose ratio is
         >= :data:`_USER_FUZZY_MATCH_THRESHOLD`.

    Returns ``None`` if no candidate clears the threshold; the caller is
    responsible for either skipping the row or auto-creating a placeholder
    user (see
    :func:`app.services.build_plan_import_service._get_or_create_inactive_user`).
    """
    requester_name = clean(requester_name)
    if not requester_name:
        return None

    stripped = _strip_parens(requester_name)
    if not stripped:
        return None

    exact = (
        session.query(User)
        .filter(User.full_name.ilike(stripped))
        .first()
    )
    if exact:
        return exact

    target = _normalize_for_match(stripped)
    if not target:
        return None

    best_user = None
    best_ratio = 0.0
    for user in session.query(User).all():
        candidate = _normalize_for_match(user.full_name or "")
        if not candidate:
            continue
        ratio = SequenceMatcher(None, target, candidate).ratio()
        if ratio > best_ratio:
            best_ratio = ratio
            best_user = user

    if best_user is not None and best_ratio >= _USER_FUZZY_MATCH_THRESHOLD:
        return best_user

    return None


def get_user_by_employee_id(session: Session, employee_id: str):
    employee_id = clean(employee_id)
    if not employee_id:
        return None

    return (
        session.query(User)
        .filter(User.employee_id == employee_id)
        .first()
    )


def get_or_create_build_plan_access(
    session: Session,
    family_form_factor: FamilyFormFactor,
    user: User,
):
    access = (
        session.query(BuildPlanAccess)
        .filter_by(
            family_form_factor_id=family_form_factor.id,
            user_id=user.id,
        )
        .first()
    )

    if access:
        return access

    access = BuildPlanAccess(
        family_form_factor_id=family_form_factor.id,
        user_id=user.id,
        access_type=AccessTypeEnum.editor,
    )
    session.add(access)
    session.flush()
    return access


def ensure_minimum_access_type(
    access: BuildPlanAccess,
    required_access_type: AccessTypeEnum,
):
    current_rank = ACCESS_LEVEL_RANK.get(access.access_type, 0)
    required_rank = ACCESS_LEVEL_RANK[required_access_type]

    if required_rank > current_rank:
        access.access_type = required_access_type


def seed_family_form_factor_access(session: Session, family_form_factor: FamilyFormFactor):
    # All authenticated users implicitly have viewer access — no row needed.
    # Only persist elevated access (editor / owner).

    owner_user = get_user_by_employee_id(session, OWNER_EMPLOYEE_ID)
    if owner_user:
        owner_access = get_or_create_build_plan_access(
            session=session,
            family_form_factor=family_form_factor,
            user=owner_user,
        )
        owner_access.access_type = AccessTypeEnum.owner
    else:
        print(f"Owner user not found by employee ID: {OWNER_EMPLOYEE_ID}")

    for employee_id in EDITOR_EMPLOYEE_IDS:
        editor_user = get_user_by_employee_id(session, employee_id)
        if not editor_user:
            print(f"Editor user not found by employee ID: {employee_id}")
            continue

        editor_access = get_or_create_build_plan_access(
            session=session,
            family_form_factor=family_form_factor,
            user=editor_user,
        )
        ensure_minimum_access_type(editor_access, AccessTypeEnum.editor)


def _iter_sample_rows(samples):
    """Normalize sample entries to ``(name, value, is_bold)`` tuples for the
    legacy "Others/bold" import path. Accepts either the new dict shape
    emitted by :func:`parse_build_plan_column` or older tuple shapes.
    """
    for entry in samples:
        if isinstance(entry, dict):
            yield entry.get("name"), entry.get("value"), bool(entry.get("is_bold"))
            continue
        if len(entry) == 3:
            name, value, is_bold = entry
        else:
            name, value = entry
            is_bold = False
        yield name, value, is_bold


def _samples_as_dicts(samples):
    """Coerce a samples list (mixed dict / tuple shapes) into the dict shape
    expected by the SUM-format importer paths. Tuple entries get
    ``formula_refs=None`` and ``row_idx=-1``.
    """
    out = []
    for entry in samples:
        if isinstance(entry, dict):
            out.append({
                "name": entry.get("name"),
                "value": entry.get("value"),
                "is_bold": bool(entry.get("is_bold")),
                "formula_refs": entry.get("formula_refs"),
                "row_idx": entry.get("row_idx", -1),
            })
            continue
        if len(entry) == 3:
            name, value, is_bold = entry
        else:
            name, value = entry
            is_bold = False
        out.append({
            "name": name,
            "value": value,
            "is_bold": bool(is_bold),
            "formula_refs": None,
            "row_idx": -1,
        })
    return out


def _samples_use_sum_format(samples) -> bool:
    """Detect whether a column's sample rows use the SUM-formula format.

    A column is treated as SUM-format if **any** sample row's value cell
    is a ``=SUM(...)`` formula referencing other rows in the same column.
    The two formats never co-exist within a single column (one PM owns
    each column), so a single positive signal is enough.
    """
    for entry in samples:
        if isinstance(entry, dict):
            refs = entry.get("formula_refs")
            if refs:
                return True
    return False


_WW_PATTERN = re.compile(r"WW\s*(\d{2})\s*(\d{2})", re.IGNORECASE)
_REV_PATTERN = re.compile(r"rev\s*0*(\d+)", re.IGNORECASE)


def parse_filename_metadata(file_path):
    """Return (work_week, work_year, file_revision) parsed from filename.

    Any element may be ``None`` when the filename does not encode it.
    Examples: ``LzP Build Plan WW1626 rev1.xlsx`` -> (16, 2026, 1).
    """
    name = Path(file_path).stem
    week = year = revision = None

    ww = _WW_PATTERN.search(name)
    if ww:
        week = int(ww.group(1))
        year = 2000 + int(ww.group(2))

    rev = _REV_PATTERN.search(name)
    if rev:
        revision = int(rev.group(1))

    return week, year, revision


def get_or_create_build_desc(session, support_activity, description):
    return get_or_create(
        session,
        BuildPlanBuildDesc,
        support_activity_id=support_activity.id,
        description=clean(description) or "N/A",
    )


def get_or_create_support_activity(session, value):
    value = clean(value) or "Integration"

    return get_or_create(
        session,
        SupportActivity,
        name=value,
    )


def split_build_notes(raw_notes):
    raw_notes = clean(raw_notes)

    if not raw_notes:
        return ["N/A"]

    # Normalize full-width comma to normal comma
    raw_notes = raw_notes.replace("，", ",")

    return [
        note.strip()
        for note in re.split(r"[,;\n]+", raw_notes)
        if note.strip()
    ]


def link_build_note_to_support_activity(
    session,
    support_activity,
    build_note,
):
    return get_or_create(
        session,
        SupportActivityBuildNote,
        support_activity_id=support_activity.id,
        build_note_id=build_note.id,
    )


def link_build_note_to_build_plan(
    session,
    build_plan,
    build_note,
):
    return get_or_create(
        session,
        BuildPlanBuildNote,
        build_plan_id=build_plan.id,
        build_note_id=build_note.id,
    )


def create_build_plan(
    session,
    family_form_factor,
    build_info,
    quantities,
):
    """Find or create the canonical build plan for ``(family_form_factor,
    config_number)`` and apply the scalar fields from ``build_info`` /
    ``quantities``. Children (components, tests, build requests, warehouse
    quantities) and revision tracking are managed by the caller.

    Returns the :class:`BuildPlan` row (created or updated in place).
    """
    config_number_value = clean(build_info.get("Config Number"))

    if not config_number_value:
        raise ValueError("Missing Config Number")

    support_activity = get_or_create_support_activity(
        session,
        build_info.get("Support Activity"),
    )

    status = map_build_status(build_info.get("Status"))

    build_desc = get_or_create_build_desc(
        session,
        support_activity,
        build_info.get("Build Description"),
    )

    config_number = get_or_create(
        session,
        ConfigNumber,
        value=config_number_value,
    )

    build_plan = (
        session.query(BuildPlan)
        .filter_by(
            family_form_factor_id=family_form_factor.id,
            config_number_id=config_number.id,
        )
        .first()
    )

    if build_plan is None:
        build_plan = BuildPlan(
            family_form_factor_id=family_form_factor.id,
            config_number_id=config_number.id,
            support_activity_id=support_activity.id,
            build_description_id=build_desc.id,
            status=status,
            is_imported=True,
        )
        session.add(build_plan)
        session.flush()

    build_plan.support_activity_id = support_activity.id
    build_plan.status = status
    build_plan.build_description_id = build_desc.id

    build_plan.ta_number = build_info.get("TA Number") or "N/A"

    build_plan.product_code = build_info.get("Product Code")
    build_plan.mm_number = build_info.get("MM Number")
    build_plan.as_number = build_info.get("AS Number")
    build_plan.pba_number = build_info.get("PBA Number")

    build_plan.special_instruction = build_info.get(
        "Special Instruction"
    )

    required_quantity = safe_int(quantities.get("Required Quantity"))
    estimated_yield = safe_percent_int(quantities.get("Estimated Yield"))
    build_start_quantity = safe_int_ceil(quantities.get("Build Start Quantity"))

    build_plan.required_quantity = required_quantity
    build_plan.estimated_yield = estimated_yield
    build_plan.build_start_quantity = build_start_quantity

    # Calendar year derived from the config number (``<FamilyCode><YY><WW>``).
    derived_year = derive_year_from_config(config_number_value)
    if derived_year is not None:
        build_plan.year = derived_year

    # ISO work-week derived from the same config-number suffix.
    derived_work_week = derive_work_week_from_config(config_number_value)
    if derived_work_week is not None:
        build_plan.work_week = derived_work_week

    session.flush()

    # ==========================
    # BUILD NOTES
    # ==========================

    raw_build_notes = build_info.get("Build Notes")

    for note_text in split_build_notes(raw_build_notes):

        build_note = get_or_create(
            session,
            BuildNote,
            notes=note_text,
        )

        link_build_note_to_build_plan(
            session,
            build_plan,
            build_note,
        )

        link_build_note_to_support_activity(
            session,
            support_activity,
            build_note,
        )

    return build_plan


def get_or_create_build_plan_component(
    session: Session,
    build_plan: BuildPlan,
    component: Component,
    slot: ComponentSlot | None,
):
    query = session.query(BuildPlanComponent).filter(
        BuildPlanComponent.build_plan_id == build_plan.id,
        BuildPlanComponent.component_id == component.id,
    )

    if slot:
        query = query.filter(BuildPlanComponent.slot_id == slot.id)
    else:
        query = query.filter(BuildPlanComponent.slot_id.is_(None))

    build_plan_component = query.first()

    if build_plan_component:
        build_plan_component.is_key = True
        return build_plan_component

    build_plan_component = BuildPlanComponent(
        build_plan_id=build_plan.id,
        component_id=component.id,
        slot_id=slot.id if slot else None,
        is_key=True,
    )

    session.add(build_plan_component)
    session.flush()
    return build_plan_component


def import_key_components(
    session: Session,
    build_plan: BuildPlan,
    key_components: dict,
    family_form_factor: "FamilyFormFactor | None" = None,
):
    # Determine which family this build plan belongs to so we can record
    # (component, supplier, family) into ``component_supplier_families``.
    # We accept an explicit ``family_form_factor`` to avoid an extra query
    # in the hot path, but fall back to looking it up from ``build_plan``
    # so legacy call sites keep working.
    family_id: int | None = None
    if family_form_factor is not None:
        family_id = family_form_factor.family_id
    elif build_plan.family_form_factor_id is not None:
        ffr = (
            session.query(FamilyFormFactor)
            .filter(FamilyFormFactor.id == build_plan.family_form_factor_id)
            .first()
        )
        if ffr is not None:
            family_id = ffr.family_id

    for field, raw_value in key_components.items():
        value = clean(raw_value)
        if not value:
            continue

        # Silicon Stepping is stored in its own table — skip it here.
        if (clean(field) or "") == SILICON_STEPPING_FIELD:
            continue

        try:
            component_name, slot_code, attr = parse_component_field(field)
        except ValueError:
            continue

        if component_name not in VALID_COMPONENTS:
            continue

        if slot_code is not None and slot_code not in VALID_SLOTS:
            raise ValueError(f"Invalid slot: {slot_code}")

        component = get_or_create(session, Component, name=component_name)

        slot = None
        if slot_code:
            slot = get_or_create(
                session,
                ComponentSlot,
                component_id=component.id,
                slot_code=slot_code,
            )

        build_plan_component = get_or_create_build_plan_component(
            session=session,
            build_plan=build_plan,
            component=component,
            slot=slot,
        )

        if attr == "Supplier":
            supplier = resolve_supplier(session, value)
            if supplier is None:
                continue

            get_or_create(
                session,
                ComponentSupplier,
                component_id=component.id,
                supplier_id=supplier.id,
            )

            # Record the (component, supplier, family) triple so admins
            # can see "supplier X supplies component Y for family Z" in
            # the DB Tables admin tree.
            if family_id is not None:
                get_or_create(
                    session,
                    ComponentSupplierFamily,
                    component_id=component.id,
                    supplier_id=supplier.id,
                    family_id=family_id,
                )

            build_plan_component.supplier_id = supplier.id
            continue

        attr_def = get_or_create(
            session,
            AttributeDefinition,
            name=attr,
            defaults={"data_type": "text"},
        )

        attr_value = (
            session.query(ComponentAttributeValue)
            .filter_by(
                build_plan_component_id=build_plan_component.id,
                attribute_id=attr_def.id,
            )
            .first()
        )

        if attr_value:
            attr_value.value_text = value
        else:
            session.add(
                ComponentAttributeValue(
                    build_plan_component_id=build_plan_component.id,
                    attribute_id=attr_def.id,
                    value_text=value,
                )
            )


# ---------------------------------------------------------------------------
# Silicon stepping
# ---------------------------------------------------------------------------

def split_silicon_steppings(raw_value) -> list[tuple[str, str | None]]:
    """Split a "Silicon Stepping" cell into ``(base, suffix)`` pairs.

    Only the five base steppings (``STC``, ``A0``, ``A1``, ``B0``, ``B1``)
    are stored in :class:`SiliconStepping`; anything else attached to a base
    stepping in the cell (e.g. ``"STC FLV3"``) is captured as a ``suffix``
    on the corresponding :class:`BuildPlanSiliconStepping` link row.

    Tokens are whitespace-separated; commas / semicolons / slashes are also
    accepted as separators. The parser walks tokens left-to-right and
    associates each non-base token with the most-recently-seen base token.
    Leading non-base tokens (with no preceding base) are ignored.

    Examples
    --------
    ``"STC"``         -> ``[("STC", None)]``
    ``"STC FLV1"``    -> ``[("STC", "FLV1")]``
    ``"A0 B0"``       -> ``[("A0", None), ("B0", None)]``
    ``"STC FLV1 A0"`` -> ``[("STC", "FLV1"), ("A0", None)]``
    ``"A0 FFFF B0"``  -> ``[("A0", "FFFF"), ("B0", None)]``
    """
    value = clean(raw_value)
    if not value:
        return []
    tokens = [t.strip() for t in re.split(r"[\s,;/]+", value) if t.strip()]
    base_set = {b.upper() for b in BASE_SILICON_STEPPINGS}

    pairs: list[tuple[str, str | None]] = []
    current_base: str | None = None
    current_suffix: list[str] = []

    def _flush():
        nonlocal current_base, current_suffix
        if current_base is not None:
            suffix = " ".join(current_suffix) if current_suffix else None
            pairs.append((current_base, suffix))
        current_base = None
        current_suffix = []

    for token in tokens:
        if token.upper() in base_set:
            _flush()
            current_base = token.upper()
        elif current_base is not None:
            current_suffix.append(token)
        # else: leading non-base token with no base -> ignore
    _flush()

    # De-duplicate by (base, suffix) preserving order.
    seen: set[tuple[str, str | None]] = set()
    ordered: list[tuple[str, str | None]] = []
    for pair in pairs:
        if pair in seen:
            continue
        seen.add(pair)
        ordered.append(pair)
    return ordered


def import_silicon_steppings(
    session: Session,
    build_plan: BuildPlan,
    key_components: dict,
):
    """Replace the silicon-stepping links for ``build_plan`` with whatever
    the parsed ``Silicon Stepping`` cell contains. Each ``(base, suffix)``
    pair becomes one link row in :class:`BuildPlanSiliconStepping`; the base
    name is upserted in :class:`SiliconStepping`."""
    raw_value = None
    for field, value in key_components.items():
        if (clean(field) or "") == SILICON_STEPPING_FIELD:
            raw_value = value
            break

    pairs = split_silicon_steppings(raw_value)

    # Wipe existing links so re-imports converge to the parsed set.
    session.query(BuildPlanSiliconStepping).filter_by(
        build_plan_id=build_plan.id
    ).delete(synchronize_session=False)
    session.flush()

    for base, suffix in pairs:
        stepping = get_or_create(session, SiliconStepping, name=base)
        get_or_create(
            session,
            BuildPlanSiliconStepping,
            build_plan_id=build_plan.id,
            silicon_stepping_id=stepping.id,
            suffix=suffix,
        )


def parse_test_field(field_name: str):
    name = clean(field_name)
    if not name:
        return None, None

    parts = name.split("_")

    if len(parts) == 1:
        return name, None

    return parts[0], "_".join(parts[1:])


def get_or_create_build_plan_test(
    session: Session,
    build_plan: BuildPlan,
    test: Test,
    test_detail: TestDetail | None,
):
    query = session.query(BuildPlanTest).filter(
        BuildPlanTest.build_plan_id == build_plan.id,
        BuildPlanTest.test_id == test.id,
    )

    if test_detail:
        query = query.filter(BuildPlanTest.test_detail_id == test_detail.id)
    else:
        query = query.filter(BuildPlanTest.test_detail_id.is_(None))

    existing = query.first()

    if existing:
        return existing

    build_plan_test = BuildPlanTest(
        build_plan_id=build_plan.id,
        test_id=test.id,
        test_detail_id=test_detail.id if test_detail else None,
    )

    session.add(build_plan_test)
    session.flush()
    return build_plan_test


def import_test_section(session: Session, build_plan: BuildPlan, test_section: dict):
    for field_name, raw_value in test_section.items():
        value = clean(raw_value)
        if not value:
            continue

        test_name, parsed_detail = parse_test_field(field_name)

        if not test_name:
            continue

        test = get_or_create(session, Test, name=test_name)
        detail_value = parsed_detail or value

        test_detail = None
        if detail_value:
            test_detail = get_or_create(
                session,
                TestDetail,
                test_id=test.id,
                detail=detail_value,
            )

        get_or_create_build_plan_test(
            session=session,
            build_plan=build_plan,
            test=test,
            test_detail=test_detail,
        )


def link_build_request_to_build_plan(
    session: Session,
    build_plan: BuildPlan,
    build_request: BuildRequest,
):
    existing_link = (
        session.query(BuildPlanBuildRequest)
        .filter_by(
            build_plan_id=build_plan.id,
            build_request_id=build_request.id,
        )
        .first()
    )

    if existing_link:
        return existing_link

    link = BuildPlanBuildRequest(
        build_plan_id=build_plan.id,
        build_request_id=build_request.id,
    )

    session.add(link)
    session.flush()
    return link


def get_build_request_for_current_build_plan(
    session: Session,
    build_plan: BuildPlan,
    user_id: int,
):
    return (
        session.query(BuildRequest)
        .join(
            BuildPlanBuildRequest,
            BuildPlanBuildRequest.build_request_id == BuildRequest.id,
        )
        .filter(
            BuildPlanBuildRequest.build_plan_id == build_plan.id,
            BuildRequest.requestor_id == user_id,
        )
        .first()
    )


def get_latest_build_request_for_same_config(
    session: Session,
    build_plan: BuildPlan,
    family_form_factor_id: int,
    user_id: int,
):
    """Return the most recent BuildRequest by ``user`` for the same
    ``(family_form_factor, config_number)`` as ``build_plan``. Since the canonical
    BuildPlan is unique per (family_form_factor, config_number), "latest" is just the
    BuildRequest with the highest ``revision``."""
    return (
        session.query(BuildRequest)
        .filter(
            BuildRequest.requestor_id == user_id,
            BuildRequest.family_form_factor_id == family_form_factor_id,
            BuildRequest.config_number_id == build_plan.config_number_id,
        )
        .order_by(BuildRequest.revision.desc())
        .first()
    )


def import_build_requests(
    session: Session,
    build_plan: BuildPlan,
    family_form_factor: FamilyFormFactor,
    samples: list,
):
    # Wipe-and-recreate: each sample row is meant to produce its own
    # ``BuildRequest`` (a single requestor may legitimately appear in
    # multiple rows of the same column, handled by different recipients,
    # and we must keep those quantities separate rather than collapsing
    # them onto the last-parsed row). To stay idempotent across re-imports
    # we first drop the per-build-plan links and any orphaned BuildRequest
    # rows that were only attached to this build plan.
    linked_br_ids = [
        bid for (bid,) in
        session.query(BuildPlanBuildRequest.build_request_id)
        .filter_by(build_plan_id=build_plan.id)
        .all()
    ]
    session.query(BuildPlanBuildRequest).filter_by(
        build_plan_id=build_plan.id
    ).delete(synchronize_session=False)
    session.flush()

    if linked_br_ids:
        still_linked = {
            bid for (bid,) in
            session.query(BuildPlanBuildRequest.build_request_id)
            .filter(BuildPlanBuildRequest.build_request_id.in_(linked_br_ids))
            .all()
        }
        to_delete = [bid for bid in linked_br_ids if bid not in still_linked]
        if to_delete:
            # Null out previous_build_request_id pointers from any BR
            # (on any build plan) that would otherwise dangle.
            session.query(BuildRequest).filter(
                BuildRequest.previous_build_request_id.in_(to_delete)
            ).update(
                {BuildRequest.previous_build_request_id: None},
                synchronize_session=False,
            )
            session.query(BuildRequest).filter(
                BuildRequest.id.in_(to_delete)
            ).delete(synchronize_session=False)
            session.flush()

    # In the SUM-formula format the row that holds the formula is the
    # *recipient* (handler); their cell value is just the sum of the
    # requestor rows. Issuing a BuildRequest for the recipient would
    # double-count the quantity, so we skip those rows entirely.
    if _samples_use_sum_format(samples):
        recipient_row_indices: set[int] = set()
        for entry in _samples_as_dicts(samples):
            if entry["formula_refs"]:
                recipient_row_indices.add(entry["row_idx"])
        sample_iter = [
            (s["name"], s["value"], s["is_bold"], s["row_idx"])
            for s in _samples_as_dicts(samples)
            if s["row_idx"] not in recipient_row_indices
        ]
    else:
        recipient_row_indices = set()
        sample_iter = [
            (name, value, is_bold, -1)
            for name, value, is_bold in _iter_sample_rows(samples)
        ]

    # Track per-requestor next revision so multiple rows for the same
    # user inside this one column get sequential revisions instead of
    # colliding on the same value.
    next_revisions: dict[int, int] = {}
    last_request_per_user: dict[int, BuildRequest | None] = {}

    for raw_name, raw_quantity, _is_bold, _row_idx in sample_iter:
        raw_name = clean(raw_name)

        if not raw_name:
            continue

        if is_others_row(raw_name):
            continue

        if is_sample_group_header(raw_name):
            continue

        if is_non_user_sample_row(raw_name):
            continue

        user = get_user_by_sample_name(session, raw_name)

        if not user:
            print(f"Skipping sample row: user not found: {raw_name}")
            continue

        raw_quantity = clean(raw_quantity)

        if not raw_quantity:
            continue

        quantity = safe_int(raw_quantity.replace("(No Test)", ""))

        if quantity is None or quantity <= 0:
            continue

        if user.id not in next_revisions:
            previous_request = get_latest_build_request_for_same_config(
                session=session,
                build_plan=build_plan,
                family_form_factor_id=family_form_factor.id,
                user_id=user.id,
            )
            next_revisions[user.id] = (
                previous_request.revision + 1 if previous_request else 1
            )
            last_request_per_user[user.id] = previous_request

        previous_for_link = last_request_per_user[user.id]

        build_request = BuildRequest(
            requestor_id=user.id,
            family_form_factor_id=family_form_factor.id,
            config_number_id=build_plan.config_number_id,
            quantity=quantity,
            status=BuildRequestStatus.none,
            revision=next_revisions[user.id],
            previous_build_request_id=(
                previous_for_link.id if previous_for_link else None
            ),
        )

        session.add(build_request)
        session.flush()

        next_revisions[user.id] += 1
        last_request_per_user[user.id] = build_request

        link_build_request_to_build_plan(
            session=session,
            build_plan=build_plan,
            build_request=build_request,
        )


def import_build_plan_shippings(
    session: Session,
    build_plan: BuildPlan,
    samples: list,
):
    """Populate per-build-plan recipient<->requestor links.

    Writes rows into ``build_plan_shippings`` with the new schema:
    ``(build_plan_id, recipient_user_id, requestor_user_id, quantity)``.
    Existing rows for ``build_plan`` are deleted first so re-imports are
    idempotent.

    Two parsing modes are supported:

    * **SUM-formula format**: the sample whose value cell is a
      ``=SUM(...)`` formula is the *recipient*; each row referenced by
      the formula is a *requestor* with the row's own quantity.
    * **Bold/group-header format** (legacy): explicit ``IDC``/``PTK``
      group headers, plus per-cell bold formatting, identify the
      handler/recipient for each block of non-bold requestor rows.
    """
    # Wipe existing rows for this build plan so re-imports stay clean.
    session.query(BuildPlanShipping).filter_by(
        build_plan_id=build_plan.id
    ).delete(synchronize_session=False)
    session.flush()

    # Only keep requestors that actually have a BuildRequest for this
    # build plan. Per spec, shipment recipients (the part-getters under a
    # handler) are scoped to the build — surface only users who actually
    # filed a build request on this plan.
    #
    # NOTE: We deliberately do NOT filter the *handler* (recipient_user_id)
    # this way. In SUM-formula columns the handler's cell is just a
    # ``=SUM(...)`` of the requestor rows, so ``import_build_requests``
    # skips it and no BuildRequest is created for the handler. Filtering
    # the handler against build_requests would therefore reject every row.
    allowed_requestor_ids: set[int] = {
        uid for (uid,) in (
            session.query(BuildRequest.requestor_id)
            .join(
                BuildPlanBuildRequest,
                BuildPlanBuildRequest.build_request_id == BuildRequest.id,
            )
            .filter(BuildPlanBuildRequest.build_plan_id == build_plan.id)
            .distinct()
            .all()
        )
    }

    seen: set[tuple[int, int]] = set()

    def _add(recipient_user_id: int, requestor_user_id: int, quantity):
        if requestor_user_id not in allowed_requestor_ids:
            return
        key = (recipient_user_id, requestor_user_id)
        if key in seen:
            return
        seen.add(key)
        session.add(
            BuildPlanShipping(
                build_plan_id=build_plan.id,
                recipient_user_id=recipient_user_id,
                requestor_user_id=requestor_user_id,
                quantity=quantity,
            )
        )

    if _samples_use_sum_format(samples):
        # Index samples by their source row index so formula refs can be
        # resolved back to the corresponding sample dict.
        by_row: dict[int, dict] = {}
        for entry in samples:
            if isinstance(entry, dict):
                row_idx = entry.get("row_idx")
                if isinstance(row_idx, int) and row_idx >= 0:
                    by_row[row_idx] = entry

        for entry in samples:
            if not isinstance(entry, dict):
                continue
            refs = entry.get("formula_refs")
            if not refs:
                continue

            recipient_name = clean(entry.get("name"))
            if not recipient_name:
                continue
            recipient_user = get_user_by_sample_name(session, recipient_name)
            if not recipient_user:
                continue

            for row_idx in refs:
                requestor_entry = by_row.get(row_idx)
                if not requestor_entry:
                    continue
                requestor_name = clean(requestor_entry.get("name"))
                if not requestor_name:
                    continue
                requestor_user = get_user_by_sample_name(session, requestor_name)
                if not requestor_user:
                    continue
                if requestor_user.id == recipient_user.id:
                    continue

                cleaned_qty = clean(requestor_entry.get("value")) or ""
                quantity = safe_int(cleaned_qty.replace("(No Test)", ""))
                _add(recipient_user.id, requestor_user.id, quantity)

        session.flush()
        return

    # Legacy bold/group-header walk. A bold name marks a recipient/handler
    # that owns the following non-bold requestor rows until the next bold
    # name or an "Others"/non-user terminator row.
    current_recipient_user_id: int | None = None

    for raw_name, raw_quantity, is_bold in _iter_sample_rows(samples):
        raw_name = clean(raw_name)
        if not raw_name:
            continue

        if is_others_row(raw_name) or is_non_user_sample_row(raw_name):
            current_recipient_user_id = None
            continue

        if is_sample_group_header(raw_name):
            # Group-header rows (IDC/PTK) themselves are not users; reset
            # so the next bold name becomes the new recipient.
            current_recipient_user_id = None
            continue

        user = get_user_by_sample_name(session, raw_name)
        if not user:
            continue

        if is_bold:
            # Bold name = new recipient/handler for subsequent rows.
            current_recipient_user_id = user.id
            continue

        if current_recipient_user_id is None:
            continue

        if user.id == current_recipient_user_id:
            continue

        cleaned_qty = clean(raw_quantity) or ""
        quantity = safe_int(cleaned_qty.replace("(No Test)", ""))
        _add(current_recipient_user_id, user.id, quantity)

    session.flush()


def import_excel(file_path):
    file_path = Path(file_path)
    work_week, work_year, file_revision = parse_filename_metadata(file_path)

    if work_week is None or work_year is None or file_revision is None:
        print(
            f"Skipping '{file_path.name}': filename does not encode WW/year/rev."
        )
        return

    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    xls = pd.ExcelFile(file_path)

    # Open the workbook a second time with openpyxl so we can read per-cell
    # font formatting (boldness) for the requester column. Some build-plan
    # files do not include the "Others (per email instruction)" terminator,
    # so boldness is the only reliable cue for ending a sample group.
    try:
        wb = load_workbook(file_path, data_only=True, read_only=False)
    except Exception as exc:  # pragma: no cover - defensive
        print(f"[WARN] openpyxl could not open {file_path.name}: {exc}")
        wb = None

    # Second pass with ``data_only=False`` so we can read raw formulas
    # (e.g. ``=SUM(B12:B15)``) from sample value cells. Required for the
    # SUM-formula sample format where the recipient row's quantity is a
    # formula referencing the requestor rows.
    try:
        wb_formula = load_workbook(file_path, data_only=False, read_only=False)
    except Exception as exc:  # pragma: no cover - defensive
        print(f"[WARN] openpyxl could not open {file_path.name} (formula pass): {exc}")
        wb_formula = None

    with SessionLocal() as session:
        try:
            seed_default_warehouses(session)

            for sheet_name in xls.sheet_names:
                if sheet_name == "Shipping Info":
                    continue

                print(f"Importing sheet: {sheet_name}")

                df = pd.read_excel(
                    file_path,
                    engine='calamine',
                    sheet_name=sheet_name,
                    header=None,
                )

                df = df.dropna(how="all").dropna(axis=1, how="all")

                ws = wb[sheet_name] if (wb is not None and sheet_name in wb.sheetnames) else None
                ws_formula = (
                    wb_formula[sheet_name]
                    if (wb_formula is not None and sheet_name in wb_formula.sheetnames)
                    else None
                )

                family_code, family_name, form_factor_name = detect_family_and_form_factor(df)

                family_form_factor = get_family_form_factor(
                    session=session,
                    family_code=family_code,
                    family_name=family_name,
                    form_factor_name=form_factor_name,
                )

                seed_family_form_factor_access(
                    session=session,
                    family_form_factor=family_form_factor,
                )

                for col in range(1, df.shape[1]):
                    parsed = parse_build_plan_column(df, col, ws=ws, ws_formula=ws_formula)

                    build_info = parsed["build_info"]
                    key_components = parsed["key_components"]
                    test_section = parsed["test_section"]
                    quantities = parsed["quantities"]
                    samples = parsed["samples"]
                    warehouse_quantities = parsed["warehouse_quantities"]

                    has_any_data = any(
                        [
                            build_info,
                            key_components,
                            test_section,
                            quantities,
                            samples,
                            warehouse_quantities,
                        ]
                    )

                    if not has_any_data:
                        continue

                    config_number = clean(build_info.get("Config Number"))

                    if not config_number:
                        print(f"Skipping column {col}: no Config Number")
                        continue

                    if config_number.strip().upper() == "TBD":
                        print(f"Skipping column {col}: TBD Config Number")
                        continue

                    print(f"Processing build plan {config_number}")

                    try:
                        build_plan = create_build_plan(
                            session=session,
                            family_form_factor=family_form_factor,
                            build_info=build_info,
                            quantities=quantities,
                        )

                        import_key_components(
                            session=session,
                            build_plan=build_plan,
                            key_components=key_components,
                            family_form_factor=family_form_factor,
                        )

                        import_silicon_steppings(
                            session=session,
                            build_plan=build_plan,
                            key_components=key_components,
                        )

                        import_test_section(
                            session=session,
                            build_plan=build_plan,
                            test_section=test_section,
                        )

                        import_build_requests(
                            session=session,
                            build_plan=build_plan,
                            family_form_factor=family_form_factor,
                            samples=samples,
                        )

                        import_build_plan_shippings(
                            session=session,
                            build_plan=build_plan,
                            samples=samples,
                        )

                        import_warehouse_quantities(
                            session=session,
                            build_plan=build_plan,
                            warehouse_quantities=warehouse_quantities,
                        )
                    except Exception as exc:
                        print(
                            f"[ERROR] Failed processing build plan "
                            f"'{config_number}' (sheet='{sheet_name}', column={col}): "
                            f"{type(exc).__name__}: {exc}"
                        )
                        raise

            session.commit()
            print("Build plan import completed successfully.")

        except Exception:
            session.rollback()
            print("Build plan import failed. Database rollback completed.")
            raise


if __name__ == "__main__":
    import_excel(
        "/home/fbinalex/NPI-IDBMS/backend/data/build plan/LzP Build Plan WW1626 rev1.xlsx"
    )