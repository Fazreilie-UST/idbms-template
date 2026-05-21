"""One-shot backfill: populate ``build_plan_shippings`` from existing
``BuildPlanImportFile`` rows.

Why this script exists
----------------------
Earlier revisions of :func:`app.scripts.seed_build_plan.import_build_plan_shippings`
required the *handler* (``recipient_user_id``) to also be a build-request
requestor on the plan. In the SUM-formula sample format the handler's cell
is just ``=SUM(...)`` of the requestor rows, so ``import_build_requests``
deliberately skips it — meaning the handler never qualified and **every**
shipping link was filtered out, leaving ``build_plan_shippings`` empty.

The bug is now fixed (the filter applies to the requestor side instead),
but existing build plans already imported under the old code have no rows
in ``build_plan_shippings``. This script re-parses each stored
``BuildPlanImportFile`` and re-runs **only** :func:`import_build_plan_shippings`
for every build plan column, leaving everything else (build_info, key
components, build_requests, warehouse quantities, revisions, ...) untouched.

Usage::

    cd backend
    python -m app.scripts.backfill_build_plan_shippings           # all files
    python -m app.scripts.backfill_build_plan_shippings <file_id> # one file
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

import pandas as pd

from app.db.session import SessionLocal
from app.models.build.build_plan import BuildPlan
from app.models.build.build_plan_import_file import (
    BuildPlanImportFile,
    BuildPlanImportStatus,
)
from app.models.build.config_number import ConfigNumber
from app.scripts import seed_build_plan as sbp
from app.services.build_plan_import_service import _is_aux_sheet


def _backfill_one_file(session, import_file: BuildPlanImportFile) -> dict:
    """Re-parse one import file and refresh ``build_plan_shippings`` only.

    Returns a small stats dict for logging.
    """
    stats = {
        "file_id": import_file.id,
        "filename": import_file.original_filename,
        "plans_refreshed": 0,
        "plans_missing": 0,
        "columns_skipped": 0,
    }

    file_path = Path(import_file.storage_path)
    if not file_path.exists():
        print(f"[file_id={import_file.id}] missing on disk: {file_path}")
        stats["error"] = "file_missing"
        return stats

    all_sheets = pd.read_excel(file_path, sheet_name=None, engine="calamine", header=None)

    try:
        wb = load_workbook(file_path, data_only=True, read_only=False)
    except Exception as exc:
        print(f"[file_id={import_file.id}] openpyxl(data_only) failed: {exc}")
        wb = None
    try:
        wb_formula = load_workbook(file_path, data_only=False, read_only=False)
    except Exception as exc:
        print(f"[file_id={import_file.id}] openpyxl(formula) failed: {exc}")
        wb_formula = None

    for sheet_name, df in all_sheets.items():
        df = df.dropna(how="all").dropna(axis=1, how="all")
        if _is_aux_sheet(sheet_name):
            continue

        try:
            family_code, family_name, form_factor_name = sbp.detect_family_and_form_factor(df)
        except ValueError as exc:
            print(f"  [sheet={sheet_name}] skipped: {exc}")
            continue

        family_form_factor = sbp.get_family_form_factor(
            session=session,
            family_code=family_code,
            family_name=family_name,
            form_factor_name=form_factor_name,
        )

        ws = wb[sheet_name] if (wb is not None and sheet_name in wb.sheetnames) else None
        ws_formula = (
            wb_formula[sheet_name]
            if (wb_formula is not None and sheet_name in wb_formula.sheetnames)
            else None
        )

        for col in range(1, df.shape[1]):
            parsed = sbp.parse_build_plan_column(df, col, ws=ws, ws_formula=ws_formula)
            build_info = parsed.get("build_info") or {}
            samples = parsed.get("samples") or []

            config_value = sbp.clean(build_info.get("Config Number"))
            if not config_value or config_value.strip().upper() == "TBD":
                stats["columns_skipped"] += 1
                continue
            if not samples:
                stats["columns_skipped"] += 1
                continue

            config = (
                session.query(ConfigNumber)
                .filter(ConfigNumber.value == config_value)
                .first()
            )
            if config is None:
                stats["plans_missing"] += 1
                continue

            build_plan = (
                session.query(BuildPlan)
                .filter_by(
                    family_form_factor_id=family_form_factor.id,
                    config_number_id=config.id,
                )
                .first()
            )
            if build_plan is None:
                stats["plans_missing"] += 1
                continue

            sbp.import_build_plan_shippings(
                session=session,
                build_plan=build_plan,
                samples=samples,
            )
            stats["plans_refreshed"] += 1

    return stats


def backfill(file_id: int | None = None) -> None:
    with SessionLocal() as session:
        # Skip audit logging — this is a maintenance backfill, not a user
        # action; bulk imports already exclude audit events the same way.
        session.info["skip_audit"] = True

        q = session.query(BuildPlanImportFile).filter(
            BuildPlanImportFile.status == BuildPlanImportStatus.success
        )
        if file_id is not None:
            q = q.filter(BuildPlanImportFile.id == file_id)
        files = q.order_by(BuildPlanImportFile.id.asc()).all()

        if not files:
            print("No successful build plan import files found.")
            return

        print(f"Backfilling build_plan_shippings from {len(files)} import file(s)...")
        totals = {"plans_refreshed": 0, "plans_missing": 0, "columns_skipped": 0}
        for import_file in files:
            print(
                f"\n=== file_id={import_file.id} "
                f"({import_file.original_filename}) ==="
            )
            stats = _backfill_one_file(session, import_file)
            for k in totals:
                totals[k] += stats.get(k, 0)
            print(f"  -> {stats}")
            session.flush()

        session.commit()
        print("\nBackfill complete:", totals)


if __name__ == "__main__":
    arg_id = int(sys.argv[1]) if len(sys.argv) > 1 else None
    backfill(arg_id)
