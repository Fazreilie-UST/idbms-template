"""Build plan bulk import service.

Wraps the parsing helpers in `app.scripts.seed_build_plan` and adds:
  * Per-file session control (so caller can persist BuildPlanImportFile status
    and the build plan rows within the same transaction).
  * Auto-creation of *inactive* placeholder users for sample requestor names
    that are not yet in the database. Tracks them in the returned summary
    so PMs can later complete the user records.
  * Filename parsing for work-week / year / file revision.

Reuses the existing helpers verbatim. Only the user lookup is replaced via a
small monkey-patch scoped to the import call.
"""

from __future__ import annotations

import re
from contextlib import contextmanager
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable

import pandas as pd
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.auth.user import User
from app.models.build.build_plan_import_file import (
    BuildPlanImportFile,
    BuildPlanImportStatus,
)
from app.models.build.build_plan_import_shipping_info import (
    BuildPlanImportShippingInfo,
)
from app.models.build.build_plan_import_si import BuildPlanImportSi
from app.models.build.config_number import ConfigNumber

from app.scripts import seed_build_plan as sbp
from app.services.build_plan_revision_service import (
    StatusRegressionError,
    process_parsed_column,
)


# ---------------------------------------------------------------------------
# Filename parsing
# ---------------------------------------------------------------------------

# Examples that should match:
#   "LzP Build Plan WW1626 rev1.xlsx"        -> ww=16, year=2026, rev=1
#   "WhP A0 Build Plan WW3325 Rev01.xlsx"    -> ww=33, year=2025, rev=1
#   "WhP B0 Build Plan (QS) WW1926 rev1.xlsx"
#   "PeP2 Build Plan ww18'26 Rev6.xlsx"      -> ww=18, year=2026, rev=6
# Allow optional separator (apostrophe / curly quotes / dash / dot / space)
# between the work-week and the two-digit year.
_WW_PATTERN = re.compile(
    r"WW\s*(\d{2})\s*[\u2018\u2019'`\-\.\s]?\s*(\d{2})",
    re.IGNORECASE,
)
_REV_PATTERN = re.compile(r"rev\s*0*(\d+)", re.IGNORECASE)


def parse_filename_metadata(filename: str) -> dict[str, int | None]:
    """Return {work_week, work_year, file_revision}. Any field may be None."""
    name = Path(filename).stem

    week = year = revision = None

    ww_match = _WW_PATTERN.search(name)
    if ww_match:
        week = int(ww_match.group(1))
        # Two-digit year -> 2000s
        year = 2000 + int(ww_match.group(2))

    rev_match = _REV_PATTERN.search(name)
    if rev_match:
        revision = int(rev_match.group(1))

    return {"work_week": week, "work_year": year, "file_revision": revision}


# ---------------------------------------------------------------------------
# Inactive user auto-creation
# ---------------------------------------------------------------------------

# Strip parenthesised annotations (e.g. "Foo (PTK)" -> "Foo") and any trailing
# free-form noise (e.g. "Foo - Solder Down only", "Foo - KEEP FOR TEST") that
# leaks in from warehouse-stash sample rows mis-classified as user names.
_PARENS_RE = re.compile(r"\([^)]*\)")
# Match a trailing " - <noise>" suffix. The noise is any variant of the
# warehouse-stash sub-buckets we know about: "keep for test", "to keep for
# test", "solder down", "solder down only", "solder-down", etc. The pattern
# is intentionally greedy on the noise side so trailing fragments like
# "- Solder Down only - extra" are stripped wholesale.
_TRAILING_NOISE_RE = re.compile(
    r"\s*[-\u2010-\u2015]\s*"
    r"(?:to\s+)?"
    r"(?:keep\s+for\s+test|solder[\s\-]*down(?:\s+only)?)"
    r"\b.*$",
    re.IGNORECASE,
)


def _sanitize_user_name(raw: str | None) -> str:
    """Normalise a raw sample-row name before storing it as a User.full_name.

    * Removes parenthesised content (e.g. ``"Ali (PTK lead)"`` -> ``"Ali"``)
    * Strips trailing warehouse-stash noise (``"- Solder Down only"`` etc.)
    * Collapses internal whitespace.
    """
    if not raw:
        return ""
    text = _PARENS_RE.sub("", raw)
    # Strip trailing noise repeatedly in case it appears twice.
    while True:
        new_text = _TRAILING_NOISE_RE.sub("", text)
        if new_text == text:
            break
        text = new_text
    return re.sub(r"\s+", " ", text).strip()


def _get_or_create_inactive_user(session: Session, full_name: str) -> User | None:
    """Find user by full_name (case-insensitive). If missing, create an
    INACTIVE placeholder user (no email, no password) so the sample row can
    still be linked. PM can later fill out the real user details."""
    name = _sanitize_user_name(full_name)
    if not name:
        return None

    user = (
        session.query(User)
        .filter(User.full_name.ilike(name))
        .first()
    )
    if user:
        return user

    user = User(
        full_name=name,
        is_active=False,
        can_login=False,
    )
    session.add(user)
    session.flush()
    return user


@contextmanager
def _patched_user_lookup(session: Session, created_names: list[str]):
    """Monkey-patch seed_build_plan.get_user_by_sample_name for the duration
    of an import so missing users are auto-created (inactive placeholder)
    and tracked in `created_names`."""
    original = sbp.get_user_by_sample_name

    # Per-import cache: full_name (lower-cased) -> User instance. The Excel
    # files we ingest reference the same handful of users dozens of times per
    # file, so memoizing the lookup turns N SELECTs into 1 per distinct name.
    user_cache: dict[str, User | None] = {}

    def patched(_session: Session, requester_name: str):
        cleaned = sbp.clean(requester_name)
        if not cleaned:
            return None

        # Sanitize BEFORE any lookup so parenthesised annotations and
        # trailing warehouse-stash noise ("- Solder Down only", "- KEEP
        # FOR TEST", ...) never leak into either the fuzzy match or a
        # newly-created placeholder user. The original lookup is fuzzy
        # and would otherwise happily match a previously-imported user
        # whose full_name still carries the noise.
        sanitized = _sanitize_user_name(cleaned)
        if not sanitized:
            return None

        cache_key = sanitized.lower()
        if cache_key in user_cache:
            return user_cache[cache_key]

        existing = original(_session, sanitized)
        if existing:
            # Heal any legacy rows whose stored full_name still carries
            # the parenthesised annotation or trailing noise so future
            # lookups (and the UI) see the clean canonical name.
            cleaned_existing = _sanitize_user_name(existing.full_name or "")
            if cleaned_existing and cleaned_existing != (existing.full_name or "").strip():
                existing.full_name = cleaned_existing
                _session.flush()
            user_cache[cache_key] = existing
            return existing

        # Skip non-user rows (Others, CNB5, group headers, ...).
        if sbp.is_others_row(sanitized) or sbp.is_non_user_sample_row(sanitized) \
                or sbp.is_sample_group_header(sanitized):
            user_cache[cache_key] = None
            return None

        user = _get_or_create_inactive_user(_session, sanitized)
        if user is not None and sanitized not in created_names:
            created_names.append(sanitized)
        user_cache[cache_key] = user
        return user

    sbp.get_user_by_sample_name = patched
    try:
        yield
    finally:
        sbp.get_user_by_sample_name = original


@contextmanager
def _patched_get_or_create(session: Session):
    """Wrap ``seed_build_plan.get_or_create`` with a per-import in-memory cache.

    The import pipeline calls ``get_or_create`` thousands of times per file
    for the same reference rows (warehouses, components, suppliers, attribute
    definitions, tests, support activities, build descriptions, build notes,
    config numbers, ...). Each call issues a fresh ``SELECT``; for 20 files
    that's the dominant DB cost.

    Once we resolve a row inside a single import we keep the ORM instance in a
    dict keyed by ``(model_name, sorted(kwargs.items()))`` and short-circuit
    subsequent lookups. The cache lives only for the duration of one file so
    objects do not get reused across the per-file commit boundary.
    """
    original = sbp.get_or_create
    cache: dict[tuple, Any] = {}

    def cached(_session: Session, model, defaults=None, **kwargs):
        try:
            key = (model.__name__, tuple(sorted(kwargs.items())))
        except TypeError:
            # Unhashable kwargs (rare) — fall back to the uncached path.
            return original(_session, model, defaults=defaults, **kwargs)

        cached_obj = cache.get(key)
        if cached_obj is not None:
            return cached_obj

        obj = original(_session, model, defaults=defaults, **kwargs)
        cache[key] = obj
        return obj

    sbp.get_or_create = cached
    try:
        yield
    finally:
        sbp.get_or_create = original


# ---------------------------------------------------------------------------
# Auxiliary sheets: "Shipping Info" and "Si"
# ---------------------------------------------------------------------------

# Sheet names we route to dedicated parsers rather than the build-plan column
# scanner. Matched case-insensitively against the stripped sheet name.
_SHIPPING_INFO_SHEET = "shipping info"
_SI_SHEET = "si"

_AUX_SHEET_NAMES = {_SHIPPING_INFO_SHEET, _SI_SHEET}


def _is_aux_sheet(sheet_name: str) -> bool:
    return sheet_name.strip().lower() in _AUX_SHEET_NAMES


def _cell_str(value: Any) -> str:
    """Return a trimmed string for a cell, or '' for NaN/None."""
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    text = str(value).strip()
    return text


def _cell_int(value: Any) -> int | None:
    """Coerce a cell to int when possible, else None.

    Accepts plain ints, floats that round-trip to int, and digit strings
    (optionally with surrounding whitespace or a trailing decimal zero).
    """
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if float(value).is_integer():
            return int(value)
        return int(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return int(text)
    except ValueError:
        try:
            f = float(text)
            return int(f)
        except ValueError:
            return None


def _normalize_header(value: Any) -> str:
    """Lower-case + collapse non-alphanumeric runs so header lookup tolerates
    minor variations ("Si lot numbers" vs "Si Lot Numbers" vs "Si  lot-numbers")."""
    text = _cell_str(value).lower()
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def _parse_shipping_info_sheet(df: pd.DataFrame) -> list[dict[str, Any]]:
    """Extract Responsibility / Name / Address rows from a "Shipping Info"
    sheet. The header row is not necessarily the first row, so we scan from
    the top until we find a row containing all three labels and use its
    column positions for the rest of the sheet.
    """
    if df.empty:
        return []

    target_labels = {"responsibility", "name", "address"}
    header_row_idx: int | None = None
    columns: dict[str, int] = {}

    for row_idx in range(df.shape[0]):
        seen: dict[str, int] = {}
        for col_idx in range(df.shape[1]):
            label = _normalize_header(df.iat[row_idx, col_idx])
            if label in target_labels and label not in seen:
                seen[label] = col_idx
        if target_labels.issubset(seen):
            header_row_idx = row_idx
            columns = seen
            break

    if header_row_idx is None:
        return []

    rows: list[dict[str, Any]] = []
    for row_idx in range(header_row_idx + 1, df.shape[0]):
        responsibility = _cell_str(df.iat[row_idx, columns["responsibility"]])
        name = _cell_str(df.iat[row_idx, columns["name"]])
        address = _cell_str(df.iat[row_idx, columns["address"]])
        if not (responsibility or name or address):
            continue
        rows.append(
            {
                "row_index": row_idx,
                "responsibility": responsibility or None,
                "name": name or None,
                "address": address or None,
            }
        )
    return rows


# Si sheet column aliases. Keys are normalized header strings (see
# ``_normalize_header``); values are the model column names.
_SI_COLUMN_ALIASES = {
    "si description": "si_description",
    "si lot numbers": "si_lot_numbers",
    "si lot number": "si_lot_numbers",
    "class test rev": "class_test_rev",
    "request qty": "request_qty",
    "request dock date": "request_dock_date",
    "commit qty": "commit_qty",
    "commit dock date": "commit_dock_date",
    "actual qty": "actual_qty",
    "actual dock date": "actual_dock_date",
    "comments": "comments",
    "comment": "comments",
}

_SI_INT_COLUMNS = {"request_qty", "commit_qty", "actual_qty"}


def _parse_si_sheet(df: pd.DataFrame) -> list[dict[str, Any]]:
    """Extract one dict per Si lot row. Header is expected to be the first
    non-empty row (already guaranteed by the caller's ``dropna(how='all')``).
    Unknown columns are ignored; missing columns map to ``None``.
    """
    if df.empty or df.shape[0] < 2:
        return []

    column_map: dict[str, int] = {}
    for col_idx in range(df.shape[1]):
        normalized = _normalize_header(df.iat[0, col_idx])
        field = _SI_COLUMN_ALIASES.get(normalized)
        if field and field not in column_map:
            column_map[field] = col_idx

    if not column_map:
        return []

    rows: list[dict[str, Any]] = []
    for row_idx in range(1, df.shape[0]):
        record: dict[str, Any] = {"row_index": row_idx - 1}
        any_value = False
        for field, col_idx in column_map.items():
            raw = df.iat[row_idx, col_idx]
            if field in _SI_INT_COLUMNS:
                value = _cell_int(raw)
                if value is not None:
                    any_value = True
                record[field] = value
            else:
                text = _cell_str(raw)
                if text:
                    any_value = True
                record[field] = text or None
        if any_value:
            rows.append(record)
    return rows


def _persist_shipping_info(
    session: Session,
    import_file: BuildPlanImportFile,
    rows: list[dict[str, Any]],
) -> int:
    """Replace shipping_info rows for this import file. Returns rows kept
    after de-duplicating on (responsibility, name, address)."""
    session.query(BuildPlanImportShippingInfo).filter(
        BuildPlanImportShippingInfo.import_file_id == import_file.id
    ).delete(synchronize_session=False)

    seen: set[tuple[str | None, str | None, str | None]] = set()
    payload: list[dict[str, Any]] = []
    for row in rows:
        key = (row["responsibility"], row["name"], row["address"])
        if key in seen:
            continue
        seen.add(key)
        payload.append(
            {
                "import_file_id": import_file.id,
                "row_index": row.get("row_index"),
                "responsibility": row["responsibility"],
                "name": row["name"],
                "address": row["address"],
            }
        )
    if payload:
        session.bulk_insert_mappings(BuildPlanImportShippingInfo, payload)
    return len(payload)


def _persist_si_rows(
    session: Session,
    import_file: BuildPlanImportFile,
    rows: list[dict[str, Any]],
) -> int:
    """Replace si_rows for this import file. Duplicates are allowed (per
    user requirement) so every parsed row is inserted."""
    session.query(BuildPlanImportSi).filter(
        BuildPlanImportSi.import_file_id == import_file.id
    ).delete(synchronize_session=False)

    if not rows:
        return 0

    payload = [
        {
            "import_file_id": import_file.id,
            "row_index": row.get("row_index"),
            "si_description": row.get("si_description"),
            "si_lot_numbers": row.get("si_lot_numbers"),
            "class_test_rev": row.get("class_test_rev"),
            "request_qty": row.get("request_qty"),
            "request_dock_date": row.get("request_dock_date"),
            "commit_qty": row.get("commit_qty"),
            "commit_dock_date": row.get("commit_dock_date"),
            "actual_qty": row.get("actual_qty"),
            "actual_dock_date": row.get("actual_dock_date"),
            "comments": row.get("comments"),
        }
        for row in rows
    ]
    session.bulk_insert_mappings(BuildPlanImportSi, payload)
    return len(payload)


# ---------------------------------------------------------------------------
# Per-file processor (mirrors sbp.import_excel but uses caller's session)
# ---------------------------------------------------------------------------

def _process_file(
    session: Session,
    file_path: Path,
    import_file: BuildPlanImportFile,
    progress_cb: Callable[[str, dict[str, Any]], None] | None = None,
) -> dict[str, Any]:
    """Parse and import a single build plan Excel file using the provided
    session. Caller is responsible for commit / rollback.

    Per-column work is delegated to
    :func:`app.services.build_plan_revision_service.process_parsed_column`,
    which decides whether the column produces a new revision, an inserted
    mid-history revision, or just a no-change touch.

    Returns a summary dict suitable for storing on BuildPlanImportFile.summary.
    """
    sbp.seed_default_warehouses(session)

    summary: dict[str, Any] = {
        "work_week": import_file.work_week,
        "work_year": import_file.work_year,
        "file_revision": import_file.file_revision,
        "sheets_processed": 0,
        "sheets_skipped": [],
        "new_build_plans": 0,
        "revisions_created": 0,
        "revisions_inserted_midstream": 0,
        "no_change_touches": 0,
        "status_errors": [],
        "columns_skipped": 0,
        "unrecorded_users": [],
        "warnings": [],
        "shipping_info_rows": 0,
        "si_rows": 0,
    }

    # Read every sheet in a single call. ``sheet_name=None`` returns a
    # ``dict[str, DataFrame]`` and only parses the xlsx archive once, which
    # is dramatically cheaper than the previous one-call-per-sheet pattern
    # (each ``pd.read_excel`` reopens and re-decompresses the workbook).
    all_sheets = pd.read_excel(
        file_path,
        engine="calamine",
        sheet_name=None,
        header=None,
    )

    loaded_sheets: list[tuple[str, pd.DataFrame]] = []
    aux_sheets: list[tuple[str, pd.DataFrame]] = []
    total_columns = 0
    for sheet_name, df in all_sheets.items():
        df = df.dropna(how="all").dropna(axis=1, how="all")
        if _is_aux_sheet(sheet_name):
            aux_sheets.append((sheet_name, df))
            continue
        loaded_sheets.append((sheet_name, df))
        if df.shape[1] > 1:
            total_columns += df.shape[1] - 1

    # Free the original dict so the per-sheet DataFrames are the only live
    # references and can be released as soon as their loop iteration ends.
    all_sheets = None

    # Open the workbook twice with openpyxl: once with cached values (for
    # font/bold metadata) and once with raw formulas (so we can detect the
    # SUM-formula sample format — recipient row = SUM of requestor rows).
    try:
        wb = load_workbook(file_path, data_only=True, read_only=False)
    except Exception as exc:  # pragma: no cover - defensive
        summary["warnings"].append(f"openpyxl could not open workbook: {exc}")
        wb = None
    try:
        wb_formula = load_workbook(file_path, data_only=False, read_only=False)
    except Exception as exc:  # pragma: no cover - defensive
        summary["warnings"].append(f"openpyxl could not open workbook (formula pass): {exc}")
        wb_formula = None

    if progress_cb:
        progress_cb("init", {"total": total_columns})

    unrecorded_users: list[str] = []

    with _patched_user_lookup(session, unrecorded_users), \
            _patched_get_or_create(session):
        for sheet_name, df in loaded_sheets:
            try:
                family_code, family_name, form_factor_name = sbp.detect_family_and_form_factor(df)
            except ValueError as exc:
                summary["sheets_skipped"].append(sheet_name)
                summary["warnings"].append(f"Sheet '{sheet_name}': {exc}")
                if progress_cb and df.shape[1] > 1:
                    # Account for the columns we'll never visit so the
                    # progress bar still reaches 100%.
                    skipped_cols = df.shape[1] - 1
                    summary["columns_skipped"] += skipped_cols
                    progress_cb("sheet_skipped", {"sheet": sheet_name, "columns": skipped_cols})
                continue

            family_form_factor = sbp.get_family_form_factor(
                session=session,
                family_code=family_code,
                family_name=family_name,
                form_factor_name=form_factor_name,
            )

            # The user who uploaded this import file is considered the owner
            # of every build plan it produces. Grant (or upgrade to) owner
            # access on the family_form_factor so these plans show up in their
            # "Managed by me" view. We intentionally do NOT call
            # ``seed_family_form_factor_access`` here so we don't grant blanket
            # owner/editor access to hard-coded employee IDs on every import.
            uploader = import_file.uploaded_by
            if uploader is not None:
                uploader_access = sbp.get_or_create_build_plan_access(
                    session=session,
                    family_form_factor=family_form_factor,
                    user=uploader,
                )
                sbp.ensure_minimum_access_type(
                    uploader_access, sbp.AccessTypeEnum.owner
                )

            for col in range(1, df.shape[1]):
                ws = wb[sheet_name] if (wb is not None and sheet_name in wb.sheetnames) else None
                ws_formula = (
                    wb_formula[sheet_name]
                    if (wb_formula is not None and sheet_name in wb_formula.sheetnames)
                    else None
                )
                parsed = sbp.parse_build_plan_column(df, col, ws=ws, ws_formula=ws_formula)

                build_info = parsed["build_info"]
                key_components = parsed["key_components"]
                test_section = parsed["test_section"]
                quantities = parsed["quantities"]
                samples = parsed["samples"]
                warehouse_quantities = parsed["warehouse_quantities"]

                if not any([build_info, key_components, test_section,
                            quantities, samples, warehouse_quantities]):
                    summary["columns_skipped"] += 1
                    if progress_cb:
                        progress_cb("plan_skipped", {"reason": "empty"})
                    continue

                config_number_value = sbp.clean(build_info.get("Config Number"))
                if not config_number_value:
                    summary["columns_skipped"] += 1
                    if progress_cb:
                        progress_cb("plan_skipped", {"reason": "missing_config_number"})
                    continue
                if config_number_value.strip().upper() == "TBD":
                    summary["columns_skipped"] += 1
                    if progress_cb:
                        progress_cb("plan_skipped", {"reason": "tbd_config_number"})
                    continue

                config_number = sbp.get_or_create(
                    session, ConfigNumber, value=config_number_value
                )

                try:
                    process_parsed_column(
                        session,
                        import_file=import_file,
                        family_form_factor=family_form_factor,
                        config_number=config_number,
                        parsed_column=parsed,
                        summary=summary,
                        seed_helpers=sbp,
                    )
                except StatusRegressionError as exc:
                    # Surface as a hard error so the file is marked failed
                    # and no partial changes are committed.
                    summary["status_errors"].append(str(exc))
                    raise

                if progress_cb:
                    progress_cb(
                        "plan_done",
                        {
                            "config_number": config_number_value,
                            "family": family_code,
                            "form_factor": form_factor_name,
                        },
                    )

            summary["sheets_processed"] += 1

    # ------------------------------------------------------------------
    # Auxiliary sheets (Shipping Info, Si)
    # ------------------------------------------------------------------
    # Make sure the import_file row has an id before we attach children
    # (it should — process_import_file commits it first — but during tests
    # callers may pass an unflushed instance).
    if import_file.id is None:
        session.flush()

    for sheet_name, df in aux_sheets:
        normalized = sheet_name.strip().lower()
        try:
            if normalized == _SHIPPING_INFO_SHEET:
                rows = _parse_shipping_info_sheet(df)
                kept = _persist_shipping_info(session, import_file, rows)
                summary["shipping_info_rows"] = kept
                summary["sheets_processed"] += 1
            elif normalized == _SI_SHEET:
                rows = _parse_si_sheet(df)
                kept = _persist_si_rows(session, import_file, rows)
                summary["si_rows"] = kept
                summary["sheets_processed"] += 1
        except Exception as exc:  # noqa: BLE001 - never fail the whole import on aux sheets
            summary["warnings"].append(
                f"Sheet '{sheet_name}': failed to parse ({type(exc).__name__}: {exc})"
            )
            summary["sheets_skipped"].append(sheet_name)

    summary["unrecorded_users"] = unrecorded_users
    return summary


# ---------------------------------------------------------------------------
# Public entry point used by the API
# ---------------------------------------------------------------------------

def process_import_file(
    session: Session,
    import_file: BuildPlanImportFile,
    progress_cb: Callable[[str, dict[str, Any]], None] | None = None,
) -> BuildPlanImportFile:
    """Run the parsing pipeline for a stored import file row.

    Status transitions: pending/failed -> processing -> success/failed.
    On failure the build plan changes are rolled back, but the
    BuildPlanImportFile row is updated with the error.

    If ``progress_cb`` is provided it receives per-build-plan progress events
    forwarded from :func:`_process_file`.
    """
    # Bulk imports are tracked via build-plan revisions, not `audit_logs`
    # (see db/bulk_import_pseudocode.md). Mark the session so the audit
    # listeners skip the thousands of rows this pipeline writes.
    session.info["skip_audit"] = True

    import_file.status = BuildPlanImportStatus.processing
    import_file.error_message = None
    session.add(import_file)
    session.commit()
    session.refresh(import_file)

    file_path = Path(import_file.storage_path)
    if not file_path.exists():
        import_file.status = BuildPlanImportStatus.failed
        import_file.error_message = f"File not found on disk: {file_path}"
        import_file.processed_at = datetime.now(timezone.utc)
        session.commit()
        return import_file

    if (
        import_file.work_week is None
        or import_file.work_year is None
        or import_file.file_revision is None
    ):
        # The metadata regex may have improved since this row was uploaded
        # (e.g. now tolerating "ww18'26"). Re-derive from the original
        # filename on every (re)process attempt and persist the result so the
        # UI reflects the correct WW/year/rev.
        rederived = parse_filename_metadata(import_file.original_filename)
        if import_file.work_week is None:
            import_file.work_week = rederived["work_week"]
        if import_file.work_year is None:
            import_file.work_year = rederived["work_year"]
        if import_file.file_revision is None:
            import_file.file_revision = rederived["file_revision"]

    if (
        import_file.work_week is None
        or import_file.work_year is None
        or import_file.file_revision is None
    ):
        import_file.status = BuildPlanImportStatus.skipped
        import_file.error_message = (
            "Filename does not encode work-week / year / revision; "
            "cannot derive chronological build plan revisions."
        )
        import_file.processed_at = datetime.now(timezone.utc)
        session.commit()
        return import_file

    try:
        summary = _process_file(
            session,
            file_path,
            import_file=import_file,
            progress_cb=progress_cb,
        )
        import_file.status = BuildPlanImportStatus.success
        import_file.summary = summary
        import_file.processed_at = datetime.now(timezone.utc)
        session.commit()
        # New families / SKUs / support activities may have been introduced
        # during the import; drop the dashboard lookup cache so the next
        # dashboard load reflects them.
        try:
            from app.services import dashboard_service

            dashboard_service.invalidate_filter_lookups_cache()
        except Exception:  # noqa: BLE001 - cache invalidation is best-effort
            pass
    except Exception as exc:  # noqa: BLE001 - surface any parse error to the user
        session.rollback()
        # Re-attach so we can update status after rollback.
        session.refresh(import_file) if import_file in session else session.add(import_file)
        import_file.status = BuildPlanImportStatus.failed
        import_file.error_message = f"{type(exc).__name__}: {exc}"
        import_file.processed_at = datetime.now(timezone.utc)
        session.commit()

    return import_file


# ---------------------------------------------------------------------------
# Lightweight pre-count (used by the frontend to size the progress bar)
# ---------------------------------------------------------------------------

def count_build_plans_in_file(file_path: Path) -> int:
    """Return the number of build-plan columns in an Excel file.

    Counts data columns (column index >= 1 after dropping fully-empty rows /
    columns) across every sheet except the auxiliary sheets ("Shipping Info"
    and "Si") which are parsed separately. This is the same denominator used
    internally by :func:`_process_file` so the frontend bar will always reach
    exactly 100% once processing finishes.
    """
    if not file_path.exists():
        return 0
    try:
        all_sheets = pd.read_excel(
            file_path,
            engine="calamine",
            sheet_name=None,
            header=None,
        )
    except Exception:  # noqa: BLE001
        return 0

    total = 0
    for sheet_name, df in all_sheets.items():
        if _is_aux_sheet(sheet_name):
            continue
        df = df.dropna(how="all").dropna(axis=1, how="all")
        if df.shape[1] > 1:
            total += df.shape[1] - 1
    return total
